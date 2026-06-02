"""Consolidated Excel register of all saved commission invoices (for Finance income)."""
from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.commission_invoices import get_commission_invoice
from app.database import get_data_dir, get_db

log = logging.getLogger(__name__)

CONSOLIDATED_DIR = get_data_dir() / "exports" / "commission_invoices"
CONSOLIDATED_FILE = CONSOLIDATED_DIR / "GBINC_Commission_Invoices_Consolidated.xlsx"

_HDR_FONT = Font(bold=True, color="FFFFFF")
_HDR_FILL = PatternFill("solid", fgColor="1A5632")
_MONEY_FMT = "#,##0.00"
_DATE_FMT = "dd-mmm-yyyy"
_PCT_FMT = "0.00%"


def _float(val: Any) -> float:
    if val is None or val == "":
        return 0.0
    try:
        return float(str(val).replace(",", "").strip())
    except (TypeError, ValueError):
        return 0.0


def _excel_date(val: str) -> datetime | str:
    if not val:
        return ""
    s = str(val).strip()[:10]
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return val


def _write_headers(ws, headers: list[str]) -> None:
    for col, title in enumerate(headers, 1):
        cell = ws.cell(1, col, title)
        cell.font = _HDR_FONT
        cell.fill = _HDR_FILL
    ws.freeze_panes = "A2"


def _autosize(ws, max_col: int, last_row: int) -> None:
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        best = 10
        for row in range(1, min(last_row, 200) + 1):
            v = ws.cell(row, col).value
            if v is not None:
                best = max(best, min(len(str(v)) + 2, 48))
        ws.column_dimensions[letter].width = best


def _all_ci_ids() -> list[int]:
    with get_db() as conn:
        rows = conn.execute(
            """
            SELECT id FROM commission_invoices
            ORDER BY COALESCE(invoice_date, '') DESC, id DESC
            """
        ).fetchall()
    return [int(r["id"]) for r in rows]


def rebuild_consolidated_commission_workbook() -> Path:
    """Rebuild master workbook from all commission invoices in the database."""
    CONSOLIDATED_DIR.mkdir(parents=True, exist_ok=True)

    line_headers = [
        "GBInc Invoice #",
        "Invoice Date",
        "Notice Date",
        "Status",
        "End Customer",
        "GBL Invoice",
        "Product",
        "Quantity (MT)",
        "FOB Value (USD)",
        "Commission %",
        "Commission (USD)",
        "Shipment Date",
        "Delivery Port",
        "Bill To",
        "CI Line Total (USD)",
        "CI ID",
        "Last Updated",
    ]
    summary_headers = [
        "GBInc Invoice #",
        "Invoice Date",
        "Notice Date",
        "Status",
        "Bill To",
        "Delivery Port",
        "Line Count",
        "Total Commission (USD)",
        "Amount in Words",
        "Customer Order #",
        "CI ID",
        "Last Updated",
    ]

    wb = openpyxl.Workbook()
    ws_lines = wb.active
    ws_lines.title = "Commission Lines"
    _write_headers(ws_lines, line_headers)

    ws_sum = wb.create_sheet("Invoice Summary")
    _write_headers(ws_sum, summary_headers)

    line_row = 2
    sum_row = 2
    grand_commission = 0.0

    for ci_id in _all_ci_ids():
        ci = get_commission_invoice(ci_id)
        if not ci:
            continue
        lines = ci.get("line_items") or []
        ci_total = _float(ci.get("total_commission"))
        grand_commission += ci_total
        updated = (ci.get("updated_at") or ci.get("created_at") or "")[:19]

        ws_sum.cell(sum_row, 1, ci.get("invoice_number") or "")
        ws_sum.cell(sum_row, 2, _excel_date(ci.get("invoice_date") or ""))
        ws_sum.cell(sum_row, 3, _excel_date(ci.get("notice_date") or ""))
        ws_sum.cell(sum_row, 4, ci.get("status") or "")
        ws_sum.cell(sum_row, 5, ci.get("bill_to_name") or "")
        ws_sum.cell(sum_row, 6, ci.get("delivery_port") or "")
        ws_sum.cell(sum_row, 7, len(lines))
        c = ws_sum.cell(sum_row, 8, ci_total)
        c.number_format = _MONEY_FMT
        ws_sum.cell(sum_row, 9, ci.get("amount_in_words") or "")
        ws_sum.cell(sum_row, 10, ci.get("customer_order_no") or "")
        ws_sum.cell(sum_row, 11, ci_id)
        ws_sum.cell(sum_row, 12, updated)
        sum_row += 1

        if not lines:
            lines = [{}]

        for line in lines:
            comm = _float(line.get("commission_value"))
            rate = _float(line.get("commission_rate"))
            ws_lines.cell(line_row, 1, ci.get("invoice_number") or "")
            ws_lines.cell(line_row, 2, _excel_date(ci.get("invoice_date") or ""))
            ws_lines.cell(line_row, 3, _excel_date(ci.get("notice_date") or ""))
            ws_lines.cell(line_row, 4, ci.get("status") or "")
            ws_lines.cell(line_row, 5, line.get("end_customer") or "")
            ws_lines.cell(line_row, 6, line.get("gbl_invoice_number") or "")
            ws_lines.cell(line_row, 7, line.get("product_description") or "")
            q = ws_lines.cell(line_row, 8, _float(line.get("quantity")) or None)
            q.number_format = "0.00"
            fob = ws_lines.cell(line_row, 9, _float(line.get("fob_value")) or None)
            fob.number_format = _MONEY_FMT
            if rate:
                pct_val = rate / 100 if rate > 1 else rate
                pct = ws_lines.cell(line_row, 10, pct_val)
                pct.number_format = _PCT_FMT
            comm_c = ws_lines.cell(line_row, 11, comm or None)
            comm_c.number_format = _MONEY_FMT
            ship = (line.get("shipment_date") or ci.get("shipment_date") or "")[:10]
            ws_lines.cell(line_row, 12, _excel_date(ship))
            ws_lines.cell(line_row, 13, ci.get("delivery_port") or "")
            ws_lines.cell(line_row, 14, ci.get("bill_to_name") or "")
            tot = ws_lines.cell(line_row, 15, ci_total)
            tot.number_format = _MONEY_FMT
            ws_lines.cell(line_row, 16, ci_id)
            ws_lines.cell(line_row, 17, updated)
            line_row += 1

    meta = wb.create_sheet("Read Me")
    meta["A1"] = "GBINC Commission Invoices — consolidated register"
    meta["A2"] = (
        "Auto-updated when commission invoices are saved in Generate. "
        "Use for Finance → Commission Income / income entries."
    )
    meta["A3"] = f"Last rebuilt: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
    meta["A4"] = f"Invoice count: {sum_row - 2}"
    meta["A5"] = f"Line count: {line_row - 2}"
    meta["A6"] = "Grand total commission (USD):"
    meta["B6"] = grand_commission
    meta["B6"].number_format = _MONEY_FMT

    _autosize(ws_lines, len(line_headers), max(line_row - 1, 1))
    _autosize(ws_sum, len(summary_headers), max(sum_row - 1, 1))

    wb.save(CONSOLIDATED_FILE)
    return CONSOLIDATED_FILE


def read_consolidated_commission_workbook() -> tuple[bytes, str]:
    path = rebuild_consolidated_commission_workbook()
    return path.read_bytes(), path.name


def refresh_consolidated_commission_workbook() -> None:
    """Safe wrapper for hooks after CI save/delete."""
    try:
        rebuild_consolidated_commission_workbook()
    except Exception:
        log.exception("Failed to rebuild consolidated commission invoice workbook")
