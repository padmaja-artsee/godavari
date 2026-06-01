"""Excel data-request template: Tracking + Commission Details."""
from __future__ import annotations

import io
import os
import re
from copy import copy
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.ci_data_fill import (
    company_label_from_deals,
    deal_to_commission_row,
    deal_to_tracking_row,
    period_label_from_filters,
    product_label_from_deals,
)

TEMPLATE_NAME = "GBInc_Commercial_Shipping_Data_Request.xlsx"
PREFILLED_TEMPLATE_NAME = "GBInc_DataRequest_Prefilled.xlsx"

_DATE_FMT = "dd-mmm-yyyy"  # Excel display: 22-May-2026
_DATE_EXAMPLE = "22-May-2026"
_TRACK_DATE_COLS = {4, 8, 11, 14}
_TRACK_QTY_COL = 5
_TRACK_INVOICE_COL = 7  # GBL Invoice — Commission Details col D links here
_COMM_DATE_COL = 5

_YELLOW = PatternFill("solid", fgColor="FFFDE7")
_PREFILL = PatternFill("solid", fgColor="E8F5E9")

_FONT_TITLE = Font(name="Palatino Linotype", size=16, bold=True, italic=True)
_FONT_LABEL = Font(name="Calibri", size=11, bold=True)
_FONT_BODY = Font(name="Calibri", size=11)
_FONT_SMALL = Font(name="Calibri", size=10, italic=True, color="444444")
_FONT_BOLD = Font(name="Calibri", size=11, bold=True)

_TRACK_HDR_ROW = 1
_TRACK_FIRST_DATA = 2
_TRACK_DATA_ROWS_DEFAULT = 20
_TRACK_COLS = 14

_COMM_HDR_ROW = 3
_COMM_FIRST_DATA = 4
_COMM_DATA_ROWS_DEFAULT = 15
_COMM_INPUT_COLS = range(3, 15)
_COMM_FORMULA_COLS = {4, 11, 14, 15}  # D links to Tracking GBL invoice col; K/N/O calculated


def _template_path() -> Path:
    return Path(__file__).resolve().parent / "assets" / TEMPLATE_NAME


def _tracking_template_path() -> Path:
    bundle = os.environ.get("LEADS_BUNDLE_BASE")
    if bundle:
        p = Path(bundle) / "app" / "assets" / "tracking_template.xlsx"
        if p.exists():
            return p
    return Path(__file__).resolve().parent / "assets" / "tracking_template.xlsx"


def _commission_details_template_path() -> Path:
    bundle = os.environ.get("LEADS_BUNDLE_BASE")
    if bundle:
        p = Path(bundle) / "app" / "assets" / "commission_details_template.xlsx"
        if p.exists():
            return p
    return Path(__file__).resolve().parent / "assets" / "commission_details_template.xlsx"


def _set(ws, row: int, col: int, value, *, fill=None, font=None, align=None, border=None, fmt=None):
    cell = ws.cell(row, col, value)
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    if border:
        cell.border = border
    if fmt:
        cell.number_format = fmt
    return cell


def _apply_date_validation(ws, ranges: list[str]) -> None:
    dv = DataValidation(
        type="date",
        operator="between",
        formula1="DATE(1900,1,1)",
        formula2="DATE(2100,12,31)",
        allow_blank=True,
        showInputMessage=True,
        promptTitle="Date",
        prompt=f"Type {_DATE_EXAMPLE} (day, 3-letter month, year).",
    )
    for ref in ranges:
        dv.add(ref)
    ws.add_data_validation(dv)


def _lock_date_columns(ws, cols: set[int], first_row: int, last_row: int) -> None:
    """Force dd-mmm-yyyy display (e.g. 22-May-2026) on every date cell in the data range."""
    for r in range(first_row, last_row + 1):
        for c in cols:
            ws.cell(r, c).number_format = _DATE_FMT


def _copy_sheet(src_ws, wb: openpyxl.Workbook, title: str):
    dst = wb.create_sheet(title)
    for row in src_ws.iter_rows():
        for cell in row:
            if isinstance(cell, openpyxl.cell.cell.MergedCell):
                continue
            new = dst.cell(cell.row, cell.column, cell.value)
            if cell.has_style:
                new.font = copy(cell.font)
                new.fill = copy(cell.fill)
                new.border = copy(cell.border)
                new.alignment = copy(cell.alignment)
                new.number_format = cell.number_format
    for merged in src_ws.merged_cells.ranges:
        dst.merge_cells(str(merged))
    for col, dim in src_ws.column_dimensions.items():
        if dim.width:
            dst.column_dimensions[col].width = dim.width
    for row_idx, dim in src_ws.row_dimensions.items():
        if dim.height:
            dst.row_dimensions[row_idx].height = dim.height
    return dst


def _build_instructions(wb: openpyxl.Workbook) -> None:
    ws = wb.active
    ws.title = "Instructions"
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 80

    _set(ws, 1, 1, "Tracking & Commission Data Request", font=_FONT_TITLE)
    ws.merge_cells("A1:B1")

    lines = [
        "",
        "Purpose",
        "Fill yellow cells and return this workbook. We use it for monthly commission schedules",
        "and to generate Commercial Invoices (one invoice per company / product / month).",
        "",
        "Sheets",
        "• Tracking — logistics: company, PO, GBL invoice, container, vessel, ETD/ETA, destination.",
        "• Commission Details — financial: sail month, ship-to, qty, rate, ocean freight, insurance, FOB, commission.",
        "",
        "Commission Details columns",
        "SR.NO · SAIL MONTH · INVOICE NO. · VESSEL SAIL DATE · SHIP TO PARTY · PORT OF DISCHARGE ·",
        "QTY · CURRENCY · RATE PER MT · VALUE · OCEAN FREIGHT · INSURANCE · FOB VALUE · COMMISSION",
        "(VALUE, FOB VALUE, and COMMISSION are calculated — do not overwrite formulas.)",
        "",
        "Linking the two sheets",
        "Enter the GBL Invoice on Tracking (col G). Commission Details col D pulls from Tracking automatically.",
        "",
        "Tips",
        "• Yellow = fill in. Dates display as 22-May-2026 (day, 3-letter month, year).",
        "• One row per shipment. Commission is 3% of FOB unless your contact specifies otherwise.",
        "",
        "Contact",
        "Padmaja Ganapathy · Godavari Biorefineries Inc · Princeton, NJ",
    ]
    for i, text in enumerate(lines, start=3):
        bold = text in ("Purpose", "Sheets", "Commission Details columns", "Linking the two sheets", "Tips", "Contact")
        _set(ws, i, 1, text, font=Font(name="Calibri", size=11, bold=bold) if bold else _FONT_BODY)
        if text and not bold:
            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)


def _build_tracking(wb: openpyxl.Workbook, *, data_rows: int = _TRACK_DATA_ROWS_DEFAULT) -> dict[str, int]:
    src_path = _tracking_template_path()
    if src_path.exists():
        src_wb = openpyxl.load_workbook(src_path)
        ws = _copy_sheet(src_wb.active, wb, "Tracking")
        hdr = (ws.cell(1, 2).value or "").strip().lower()
        if not hdr.startswith("company"):
            ws.insert_cols(2)
            _set(ws, 1, 2, "Company", font=Font(name="Calibri", size=11, bold=True))
    else:
        ws = wb.create_sheet("Tracking")
        headers = (
            "Sr. No", "Company", "PO number ", "PO date ", "Quantity ", "Packing ",
            "GBL Invoice ", "GBL\nInvoice date ", "Container number", "Vessel Name",
            "ETD India", "Transit time", "Destination ", "ETA ",
        )
        for c, h in enumerate(headers, start=1):
            _set(ws, 1, c, h, font=Font(bold=True))

    last_data = _TRACK_FIRST_DATA + data_rows - 1
    for i in range(data_rows):
        r = _TRACK_FIRST_DATA + i
        _set(ws, r, 1, i + 1, font=_FONT_BODY, align=Alignment(horizontal="center"))
        for c in range(2, _TRACK_COLS + 1):
            fmt = _DATE_FMT if c in _TRACK_DATE_COLS else ("#,##0.00" if c == _TRACK_QTY_COL else "General")
            _set(ws, r, c, "", fill=_YELLOW, font=_FONT_BODY, fmt=fmt)

    _apply_date_validation(ws, [
        f"{get_column_letter(c)}{_TRACK_FIRST_DATA}:{get_column_letter(c)}{last_data}"
        for c in _TRACK_DATE_COLS
    ])
    _lock_date_columns(ws, _TRACK_DATE_COLS, _TRACK_FIRST_DATA, last_data)
    ws.freeze_panes = f"A{_TRACK_FIRST_DATA}"
    return {"first_row": _TRACK_FIRST_DATA, "last_row": last_data}


def _build_commission_details(wb: openpyxl.Workbook, *, data_rows: int = _COMM_DATA_ROWS_DEFAULT) -> dict[str, int]:
    src_path = _commission_details_template_path()
    if src_path.exists():
        src_wb = openpyxl.load_workbook(src_path)
        ws = _copy_sheet(src_wb.active, wb, "Commission Details")
    else:
        ws = wb.create_sheet("Commission Details")
        ws.merge_cells("B2:P2")
        _set(ws, 2, 2, "GBINC COMMISSION TOWARDS SUPPLY OF ", font=_FONT_BOLD)
        headers = (
            "SR.NO", "SAIL  MONTH", "INVOICE NO. ", "VESSEL SAIL DATE", "SHIP TO PARTY",
            "PORT OF DISCHARGE", "QTY", "CURRENCY", "RATE PER MT", "VALUE",
            "OCEAN FRIEGHT", "INSURANCE", "FOB VALUE", "COMMISSION", "CURRENCY",
        )
        for c, h in enumerate(headers, start=2):
            _set(ws, 3, c, h, font=_FONT_BOLD)

    # Unmerge title; fixed prefix + yellow product code
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row <= 2 <= merged.max_row:
            ws.unmerge_cells(str(merged))
    _set(ws, 2, 2, "GBINC COMMISSION TOWARDS SUPPLY OF ", font=_FONT_BOLD)
    ws.merge_cells("B2:O2")
    _set(ws, 2, 16, "", fill=_YELLOW, font=_FONT_BOLD)
    ws.merge_cells("P2:P2")

    # Reporting period (company + month — for monthly archive)
    _set(ws, 1, 2, "Company:", font=_FONT_LABEL)
    ws.merge_cells("C1:E1")
    _set(ws, 1, 3, "", fill=_YELLOW, font=_FONT_BODY)
    _set(ws, 1, 6, "Month / FY:", font=_FONT_LABEL)
    ws.merge_cells("G1:I1")
    _set(ws, 1, 7, "", fill=_YELLOW, font=_FONT_BODY)
    _set(ws, 1, 10, "e.g. AIMCO · April 2026-27", font=_FONT_SMALL)

    # Remove sample TOTAL row if present at row 5
    if ws.max_row >= 5 and ws.cell(5, 2).value == "TOTAL":
        for merged in list(ws.merged_cells.ranges):
            if merged.min_row == 5 and merged.max_row == 5:
                ws.unmerge_cells(str(merged))
        ws.delete_rows(5, 1)

    # delete_rows leaves stale B:G merge ranges on data rows — clear before filling
    for merged in list(ws.merged_cells.ranges):
        if (
            merged.min_row >= _COMM_FIRST_DATA
            and merged.min_row == merged.max_row
            and merged.min_col == 2
            and merged.max_col == 7
        ):
            ws.unmerge_cells(str(merged))

    last_data = _COMM_FIRST_DATA + data_rows - 1
    sample_row = _COMM_FIRST_DATA

    for i in range(1, data_rows):
        r = _COMM_FIRST_DATA + i
        ws.insert_rows(r)
        for c in range(2, 17):
            src = ws.cell(sample_row, c)
            dst = ws.cell(r, c)
            if src.has_style:
                dst.font = copy(src.font)
                dst.border = copy(src.border)
                dst.alignment = copy(src.alignment)
                dst.number_format = _DATE_FMT if c == _COMM_DATE_COL else src.number_format

    for i in range(data_rows):
        r = _COMM_FIRST_DATA + i
        ws.cell(r, 2, i + 1)
        for c in _COMM_INPUT_COLS:
            if c in _COMM_FORMULA_COLS:
                continue
            fmt = _DATE_FMT if c == _COMM_DATE_COL else ("#,##0.00" if c in (8, 10, 12, 13) else "General")
            _set(ws, r, c, "", fill=_YELLOW, font=_FONT_BODY, fmt=fmt)
        track_r = _TRACK_FIRST_DATA + i
        ws.cell(r, 4).value = f"=Tracking!{get_column_letter(_TRACK_INVOICE_COL)}{track_r}"
        ws.cell(r, 4).number_format = "General"
        ws.cell(r, 11).value = f"=H{r}*J{r}"
        ws.cell(r, 11).number_format = "#,##0.00"
        ws.cell(r, 14).value = f"=K{r}-L{r}-M{r}"
        ws.cell(r, 14).number_format = "#,##0.00"
        ws.cell(r, 15).value = f"=N{r}*3%"
        ws.cell(r, 15).number_format = "#,##0.00"

    total_row = last_data + 1
    ws.cell(total_row, 2, "TOTAL")
    ws.cell(total_row, 2).font = _FONT_BOLD
    ws.merge_cells(start_row=total_row, start_column=2, end_row=total_row, end_column=7)
    for col, letter in ((8, "H"), (11, "K"), (14, "N"), (15, "O")):
        ws.cell(total_row, col).value = f"=SUM({letter}{_COMM_FIRST_DATA}:{letter}{last_data})"
        ws.cell(total_row, col).font = _FONT_BOLD
        ws.cell(total_row, col).number_format = "#,##0.00"

    cur_dv = DataValidation(type="list", formula1='"USD,EUR"', allow_blank=True)
    cur_dv.add(f"I{_COMM_FIRST_DATA}:I{last_data}")
    cur_dv.add(f"P{_COMM_FIRST_DATA}:P{last_data}")
    ws.add_data_validation(cur_dv)

    _apply_date_validation(ws, [f"E{_COMM_FIRST_DATA}:E{last_data}"])
    _lock_date_columns(ws, {_COMM_DATE_COL}, _COMM_FIRST_DATA, last_data)

    ws.freeze_panes = f"B{_COMM_FIRST_DATA}"
    return {
        "first_row": _COMM_FIRST_DATA,
        "last_row": last_data,
        "total_row": total_row,
        "hdr_row": _COMM_HDR_ROW,
    }


def _build_field_map(wb: openpyxl.Workbook, track: dict, comm: dict) -> None:
    map_ws = wb.create_sheet("Field Map")
    map_ws.sheet_state = "hidden"
    map_ws.append(["section", "field_key", "sheet", "cell_or_range"])
    tr, lr = track["first_row"], track["last_row"]
    cr, cl = comm["first_row"], comm["last_row"]
    rows = [
        ("tracking", "line_items", "Tracking",
         f"A{tr}:N{lr} (sr,company,po,po_date,qty,packing,gbl_invoice,gbl_date,container,vessel,etd,transit,dest,eta)"),
        ("commission", "company", "Commission Details", "C1"),
        ("commission", "period", "Commission Details", "G1"),
        ("commission", "product", "Commission Details", "P2"),
        ("commission", "line_items", "Commission Details",
         f"B{cr}:P{cl} (sr,sail_month,invoice,vessel_date,ship_to,port,qty,currency,rate,value,freight,insurance,fob,commission)"),
    ]
    for r in rows:
        map_ws.append(list(r))


def _write_cell(ws, row: int, col: int, value, *, prefilled: bool = False, is_date: bool = False):
    cell = ws.cell(row, col)
    if is_date:
        cell.number_format = _DATE_FMT
    if value is None or value == "":
        return
    cell.value = value
    if prefilled:
        cell.fill = _PREFILL


def _fill_workbook_from_deals(
    wb: openpyxl.Workbook,
    deals: list[dict],
    track_meta: dict,
    comm_meta: dict,
    *,
    company: str = "",
    period_label: str = "",
    product: str = "",
    month_hint: int = 0,
) -> None:
    if not deals:
        return
    track_ws = wb["Tracking"]
    comm_ws = wb["Commission Details"]

    co_label = company_label_from_deals(deals, company)
    prod_label = product or product_label_from_deals(deals)
    if co_label:
        comm_ws["C1"] = co_label
        comm_ws["C1"].fill = _PREFILL
    if period_label:
        comm_ws["G1"] = period_label
        comm_ws["G1"].fill = _PREFILL
    if prod_label:
        comm_ws["P2"] = prod_label
        comm_ws["P2"].fill = _PREFILL

    for i, deal in enumerate(deals):
        tr = track_meta["first_row"] + i
        cr = comm_meta["first_row"] + i
        for col, val in deal_to_tracking_row(deal, i + 1).items():
            _write_cell(track_ws, tr, col, val, prefilled=True, is_date=(col in _TRACK_DATE_COLS))
        for col, val in deal_to_commission_row(deal, i + 1, month_hint=month_hint).items():
            if col in _COMM_FORMULA_COLS:
                continue
            _write_cell(
                comm_ws, cr, col, val, prefilled=True,
                is_date=(col == _COMM_DATE_COL),
            )


def build_data_request_workbook(
    deals: list[dict] | None = None,
    *,
    company: str = "",
    period_label: str = "",
    product: str = "",
    month_hint: int = 0,
) -> openpyxl.Workbook:
    n = len(deals) if deals else 0
    track_rows = max(_TRACK_DATA_ROWS_DEFAULT, n) if n else _TRACK_DATA_ROWS_DEFAULT
    comm_rows = max(_COMM_DATA_ROWS_DEFAULT, n) if n else _COMM_DATA_ROWS_DEFAULT

    wb = openpyxl.Workbook()
    _build_instructions(wb)
    track_meta = _build_tracking(wb, data_rows=track_rows)
    comm_meta = _build_commission_details(wb, data_rows=comm_rows)
    _build_field_map(wb, track_meta, comm_meta)
    if deals:
        _fill_workbook_from_deals(
            wb, deals, track_meta, comm_meta,
            company=company, period_label=period_label, product=product,
            month_hint=month_hint,
        )
        _lock_date_columns(
            wb["Tracking"], _TRACK_DATE_COLS,
            track_meta["first_row"], track_meta["last_row"],
        )
        _lock_date_columns(
            wb["Commission Details"], {_COMM_DATE_COL},
            comm_meta["first_row"], comm_meta["last_row"],
        )
    return wb


def generate_data_request_template() -> tuple[bytes, str]:
    wb = build_data_request_workbook()
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    base = TEMPLATE_NAME.removesuffix(".xlsx")
    fname = f"{base}_{_download_date_stamp()}.xlsx"
    return buf.read(), fname


def _safe_filename_part(text: str) -> str:
    t = re.sub(r"[^\w.\- ]", "_", (text or "All").strip())
    return re.sub(r"\s+", "_", t).strip("_")[:40] or "All"


def _download_date_stamp() -> str:
    from datetime import date
    return date.today().strftime("%d-%b-%Y")


def generate_prefilled_data_request(
    deals: list[dict],
    *,
    company: str = "",
    period_label: str = "",
    product: str = "",
    month_hint: int = 0,
) -> tuple[bytes, str]:
    wb = build_data_request_workbook(
        deals,
        company=company,
        period_label=period_label,
        product=product,
        month_hint=month_hint,
    )
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    parts = [_safe_filename_part(company or company_label_from_deals(deals, ""))]
    if period_label:
        parts.append(_safe_filename_part(period_label))
    parts.append(_download_date_stamp())
    fname = f"GBInc_DataRequest_{'_'.join(parts)}.xlsx"
    return buf.read(), fname


def build_data_request_workbook_empty() -> openpyxl.Workbook:
    return build_data_request_workbook()


def write_data_request_template(path: Path | None = None) -> Path:
    path = path or _template_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = build_data_request_workbook_empty()
    wb.save(path)
    return path


if __name__ == "__main__":
    out = write_data_request_template()
    print(f"Wrote {out}")
