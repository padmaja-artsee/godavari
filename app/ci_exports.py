"""Commission Invoice Excel export — fills app/assets/commission_invoice_template.xlsx."""
from __future__ import annotations

import io
import os
import re
from copy import copy
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import PatternFill

from app.commission_invoices import _float, dollars_in_words, safe_ci_filename
from app.database import get_data_dir
from app.document_assets import signature_path
from app.po_exports import export_po_pdf_html

XLSX_DIR = get_data_dir() / "exports" / "commission_invoices" / "xlsx"
PDF_DIR = get_data_dir() / "exports" / "commission_invoices" / "pdf"

_MISSING_FILL = PatternFill("solid", fgColor="FFFDE7")  # light yellow highlight

_DATA_START = 19
_DATA_SLOTS = 8
_WORDS_ROW = 27
_NOTICE_DATA = 69


def _template_path() -> Path:
    bundle = os.environ.get("LEADS_BUNDLE_BASE")
    if bundle:
        p = Path(bundle) / "app" / "assets" / "commission_invoice_template.xlsx"
        if p.exists():
            return p
    return Path(__file__).resolve().parent / "assets" / "commission_invoice_template.xlsx"


def _excel_date(val: str) -> datetime | str:
    if not val:
        return ""
    s = str(val).strip()[:10]
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return val


def _rate_decimal(rate_pct: float) -> float:
    """Template column E uses decimal rate (0.03 for 3%)."""
    r = _float(rate_pct)
    return round(r / 100, 4) if r > 1 else r


def _copy_style(src, dst) -> None:
    if src and dst:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format


def _mark_missing(ws, coord: str, value: Any, placeholder: str = "⚠ NEEDS INPUT") -> None:
    cell = ws[coord]
    cell.value = value if value not in (None, "", 0) else placeholder
    if cell.value == placeholder:
        cell.fill = _MISSING_FILL


def _find_row_with(ws, col: int, text: str, start: int = 1) -> int | None:
    for r in range(start, ws.max_row + 1):
        v = ws.cell(r, col).value
        if v and str(v).strip().lower() == text.lower():
            return r
    return None


def _add_signature_image(ws, anchor: str) -> None:
    path = signature_path()
    if not path:
        return
    try:
        from openpyxl.drawing.image import Image as XLImage

        img = XLImage(str(path))
        img.width = 130
        img.height = 45
        ws.add_image(img, anchor)
    except Exception:
        pass


def build_ci_workbook(ci: dict[str, Any]) -> openpyxl.Workbook:
    path = _template_path()
    if not path.exists():
        raise FileNotFoundError(f"Commission invoice template not found: {path}")

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    safe_title = re.sub(r'[\\/*?:\[\]]', "-", ci.get("invoice_number") or "Commission")[:31]
    ws.title = safe_title or "Commission"

    lines = ci.get("line_items") or [{}]
    n = len(lines)
    extra = max(0, n - _DATA_SLOTS)
    if extra:
        ws.insert_rows(_WORDS_ROW, extra)

    words_row = _WORDS_ROW + extra
    notice_row = _NOTICE_DATA + extra
    total_row = _find_row_with(ws, 5, "Total") or (30 + extra)

    sample = {c: ws.cell(_DATA_START, c) for c in range(1, 8)}

    # ── Header (yellow per-invoice fields) ───────────────────────────────────
    inv_date = _excel_date(ci.get("invoice_date") or "")
    ws["F3"] = inv_date
    ws["F4"] = ci.get("invoice_number") or "⚠ NEEDS INPUT"
    if not ci.get("invoice_number"):
        ws["F4"].fill = _MISSING_FILL

    ws["A16"] = ci.get("contact_person") or "Padmaja Ganapathy"
    ws["F16"] = ci.get("payment_terms") or "PROMPT"

    products = []
    # ── Line items (rows 19+) ────────────────────────────────────────────────
    for i, li in enumerate(lines):
        r = _DATA_START + i
        end_co = (li.get("end_customer") or li.get("company") or "").strip()
        product = (li.get("product_description") or "").strip()
        gbl = (li.get("gbl_invoice_number") or "").strip()
        qty = _float(li.get("quantity"))
        rate_dec = _rate_decimal(li.get("commission_rate"))
        fob = _float(li.get("fob_value"))
        if not fob and qty and _float(li.get("unit_price")):
            fob = round(qty * _float(li.get("unit_price")), 2)
        comm = _float(li.get("commission_value"))
        if not comm and fob and rate_dec:
            comm = round(fob * rate_dec, 2)

        if product:
            products.append(product)

        mapping = [
            (1, end_co or "⚠ NEEDS INPUT"),
            (2, gbl or f"=F4"),
            (3, qty or None),
            (4, f"=F{r}/E{r}" if rate_dec else (fob or None)),
            (5, rate_dec or None),
            (6, comm or None),
            (7, product or "⚠ NEEDS INPUT"),
        ]
        for col, val in mapping:
            cell = ws.cell(r, col)
            if col == 2 and gbl:
                cell.value = gbl
            elif col == 4 and not rate_dec and fob:
                cell.value = fob
            else:
                cell.value = val
            if col in sample:
                _copy_style(sample[col], cell)
            if isinstance(val, str) and val.startswith("⚠"):
                cell.fill = _MISSING_FILL

    last_data = _DATA_START + n - 1
    total_comm = round(sum(_float(li.get("commission_value")) or (
        _float(li.get("fob_value")) * _rate_decimal(li.get("commission_rate"))
    ) for li in lines), 2)

    # Amount in words + total formulas
    words = (ci.get("amount_in_words") or "").strip() or dollars_in_words(total_comm)
    ws.cell(words_row, 1, words)
    ws.cell(words_row, 6, f'=IF(SUM(F{_DATA_START}:F{last_data})>0,SUM(F{_DATA_START}:F{last_data}),"")')
    ws.cell(total_row, 6, f"=F{words_row}")

    # ── Notice of Order section ──────────────────────────────────────────────
    notice_seller_row = 53 + extra
    notice_customer_row = 60 + extra
    delivery_row = 66 + extra
    notice_date = _excel_date(ci.get("notice_date") or ci.get("invoice_date") or "")
    ws.cell(notice_seller_row, 6, notice_date)
    first_end = (lines[0].get("end_customer") or "").strip() if lines else ""
    ws.cell(notice_customer_row, 1, first_end or "⚠ NEEDS INPUT")
    if not first_end:
        ws.cell(notice_customer_row, 1).fill = _MISSING_FILL
    ws.cell(delivery_row, 1, ci.get("contact_person") or "")
    delivery = (ci.get("delivery_port") or "").strip()
    _mark_missing(ws, f"C{delivery_row}", delivery)
    ship_cell = f"F{notice_row}"
    ship_val = _excel_date(ci.get("shipment_date") or "")
    ws[ship_cell] = ship_val if ship_val else "⚠ NEEDS INPUT"
    if not ci.get("shipment_date"):
        ws[ship_cell].fill = _MISSING_FILL

    notice_extra = max(0, n - 1)
    if notice_extra:
        ws.insert_rows(notice_row + 1, notice_extra)

    for i, li in enumerate(lines):
        nr = notice_row + i
        end_co = (li.get("end_customer") or "").strip()
        product = (li.get("product_description") or "").strip()
        qty = _float(li.get("quantity"))
        rate_dec = _rate_decimal(li.get("commission_rate"))
        cif = _float(li.get("cif_price"))

        ws.cell(nr, 1, f"=A{_DATA_START + i}")
        ws.cell(nr, 2, f"=G{_DATA_START + i}")
        if cif:
            ws.cell(nr, 3, cif)
        else:
            _mark_missing(ws, f"C{nr}", None, "⚠ NEEDS INPUT")
        ws.cell(nr, 4, f"=C{_DATA_START + i}")
        ws.cell(nr, 5, rate_dec or None)
        line_ship = _excel_date(li.get("shipment_date") or "")
        if not line_ship and i == 0:
            line_ship = _excel_date(ci.get("shipment_date") or "")
        if line_ship:
            ws.cell(nr, 6, line_ship)
        elif i == 0:
            ws.cell(nr, 6, "⚠ NEEDS INPUT")
            ws.cell(nr, 6).fill = _MISSING_FILL
        else:
            ws.cell(nr, 6, f"=F{notice_row}")

    # Document title product list in B2 area — template uses static; update if product known
    if products:
        prod_label = products[0] if len(products) == 1 else ", ".join(products[:3])
        ws["B2"] = f"GBINC COMMISSION TOWARDS SUPPLY OF {prod_label.upper()}"

    # Drop embedded images (broken refs break openpyxl save)
    if hasattr(ws, "_images"):
        ws._images = []

    sig_row = 36 + extra
    _add_signature_image(ws, f"D{sig_row}")

    return wb


def export_ci_pdf(ci: dict[str, Any], html: str) -> tuple[bytes, str] | None:
    result = export_po_pdf_html(html)
    if not result:
        return None
    pdf_bytes, _ = result
    PDF_DIR.mkdir(parents=True, exist_ok=True)
    fname = f"CI_{safe_ci_filename(ci.get('invoice_number', 'CI'))}.pdf"
    (PDF_DIR / fname).write_bytes(pdf_bytes)
    return pdf_bytes, fname


def export_ci_xlsx(ci: dict[str, Any]) -> tuple[bytes, str]:
    XLSX_DIR.mkdir(parents=True, exist_ok=True)
    fname = f"CI_{safe_ci_filename(ci.get('invoice_number', 'CI'))}.xlsx"
    path = XLSX_DIR / fname
    wb = build_ci_workbook(ci)
    wb.save(path)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read(), fname
