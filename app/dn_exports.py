"""Delivery Note — Excel export. Self-contained."""

from __future__ import annotations

import io
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_GREEN     = "1A5632"
_LT_GREEN  = "E6F4EC"
_WHITE     = "FFFFFF"
_GREY      = "F3F4F6"

_thin  = Side(style="thin",   color="A0B8A8")
_thick = Side(style="medium", color=_GREEN)

def _border(**kw):
    sides = {k: (_thick if v == "thick" else _thin if v else Side(style=None))
             for k, v in kw.items()}
    return Border(**sides)

def _fill(c): return PatternFill("solid", fgColor=c)
def _font(bold=False, size=10, color="000000", italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)
def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _w(ws, r, c, val, font=None, fill=None, align=None, border=None):
    cell = ws.cell(row=r, column=c, value=val)
    if font:   cell.font      = font
    if fill:   cell.fill      = fill
    if align:  cell.alignment = align
    if border: cell.border    = border
    return cell

def _m(ws, r1, c1, r2, c2, val=None, font=None, fill=None, align=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1, value=val)
    if font:  cell.font      = font
    if fill:  cell.fill      = fill
    if align: cell.alignment = align
    return cell

def _num(v):
    try:
        f = float(str(v or "").replace(",", ""))
        return f if f else ""
    except Exception:
        return v or ""


def export_dn_xlsx(dn: dict[str, Any]) -> tuple[bytes, str]:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Delivery Note"

    for i, w in enumerate([30, 20, 14, 14, 14, 0], 1):
        if w: ws.column_dimensions[get_column_letter(i)].width = w

    r = 1

    # Company name
    _m(ws, r, 1, r, 3, val=dn.get("company_name") or "Godavari Biorefineries B.V.",
       font=_font(bold=True, size=13, color=_GREEN), align=_align("left"))
    r += 1

    # Document title
    _m(ws, r, 1, r, 5,
       val=dn.get("document_title") or "DELIVERY NOTE CUM PACKING LIST",
       font=_font(bold=True, size=13, color=_WHITE),
       fill=_fill(_GREEN), align=_align("center"))
    r += 1

    # Reference block (cols 1-2 = bill-to, cols 4-5 = reference)
    ref_pairs = [
        ("Reference No", dn.get("reference_number") or ""),
        ("Date",         dn.get("reference_date") or ""),
        ("Delivery / Pick date", dn.get("delivery_pick_date") or ""),
        ("Order Number", dn.get("order_number") or ""),
        ("ZMPN",         dn.get("zmpn") or ""),
    ]
    bill_rows = [
        ("TO",                       dn.get("bill_to_name") or ""),
        ("",                         dn.get("bill_to_address") or ""),
        ("VAT Number",               dn.get("bill_to_vat") or ""),
        ("Delivery Contact @ Site",  dn.get("delivery_contact") or ""),
        ("Name of transporter",      dn.get("transporter_name") or ""),
    ]
    for i, ((lbl_b, val_b), (lbl_r, val_r)) in enumerate(zip(bill_rows, ref_pairs)):
        _w(ws, r, 1, lbl_b, font=_font(bold=bool(lbl_b)))
        _m(ws, r, 2, r, 3, val=val_b, font=_font())
        _w(ws, r, 4, lbl_r, font=_font(bold=True))
        _w(ws, r, 5, val_r, font=_font())
        r += 1

    r += 1

    # Delivery details
    delivery_fields = [
        ("Delivery Date requested", "delivery_date_requested"),
        ("Delivery Address",        "delivery_address"),
        ("VAT Number of Delivery address", "delivery_vat"),
        ("Slot booking reference",  "slot_booking_ref"),
        ("Delivery Time slot or Standard", "delivery_time_slot"),
    ]
    for label, key in delivery_fields:
        _w(ws, r, 1, label, font=_font(bold=True))
        _m(ws, r, 2, r, 5, val=dn.get(key) or "", font=_font())
        r += 1

    r += 1

    # Product summary
    _w(ws, r, 1, "Product", font=_font(bold=True))
    _m(ws, r, 2, r, 5, val=dn.get("product_name") or "", font=_font(bold=True))
    r += 1
    _w(ws, r, 1, "Number of Packs", font=_font(bold=True))
    _w(ws, r, 2, _num(dn.get("number_of_packs")), font=_font())
    _w(ws, r, 3, "Total Quantity",  font=_font(bold=True))
    _w(ws, r, 4, _num(dn.get("total_quantity")), font=_font())
    _w(ws, r, 5, dn.get("quantity_unit") or "Kgs", font=_font())
    r += 1
    _w(ws, r, 2, dn.get("pack_unit") or "Drums", font=_font(italic=True, color="555555"))
    r += 1

    # Packaging table header
    hdr = ["Type of Packaging", "Description", "Net weight (kg)", "Tare weight (kg)", "Gross Weight (kg)"]
    for c, h in enumerate(hdr, 1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font      = _font(bold=True, color=_WHITE)
        cell.fill      = _fill(_GREEN)
        cell.alignment = _align("center", wrap=True)
    ws.row_dimensions[r].height = 28
    r += 1

    ptype  = dn.get("packaging_type") or "DRUMS"
    pdesc  = dn.get("pack_description") or "Each"
    n      = _num(dn.get("number_of_packs")) or 0
    net_e  = _num(dn.get("net_weight_each"))  or 0
    tare_e = _num(dn.get("tare_weight_each")) or 0
    gros_e = _num(dn.get("gross_weight_each")) or 0
    t_net  = _num(dn.get("total_net_weight"))  or 0
    t_tar  = _num(dn.get("total_tare_weight")) or 0
    t_gro  = _num(dn.get("total_gross_weight")) or 0

    pack_rows = [
        (ptype,  pdesc,           net_e,  tare_e, gros_e),
        ("",     "Number of Packs", n,    n,       n),
        ("",     "Total weight",   t_net,  t_tar,   t_gro),
    ]
    for pr in pack_rows:
        for c, v in enumerate(pr, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font      = _font()
            cell.alignment = _align("right" if c >= 3 else "left")
        r += 1

    # Pallet + total gross
    _w(ws, r, 1, "Pallet wt extra", font=_font(bold=True))
    _w(ws, r, 5, _num(dn.get("pallet_weight_extra")) or "", font=_font())
    r += 1
    _w(ws, r, 1, "Total Gross weight", font=_font(bold=True))
    _w(ws, r, 5, _num(dn.get("total_gross_weight")) or "", font=_font(bold=True))
    r += 1
    r += 1

    # Batch / product details
    detail_pairs = [
        ("Batch Number",           "batch_number"),
        ("Manufacturing Date",     "manufacturing_date"),
        ("Expiry Date",            "expiry_date"),
        ("Net weight of each pack","net_weight_each"),
        ("Tare weight of each pack","tare_weight_each"),
        ("Gross weight of each pack","gross_weight_each"),
    ]
    for label, key in detail_pairs:
        _w(ws, r, 1, label, font=_font(bold=True))
        v = dn.get(key)
        _m(ws, r, 2, r, 3, val=(_num(v) if key.endswith(("_each","_each")) else (v or "")), font=_font())
        r += 1

    r += 1
    # Additional info
    add_pairs = [
        ("Number of Pallets",       "number_of_pallets"),
        ("Pallet wt extra at actual", "pallet_weight_actual"),
        ("Manufacturer",            "manufacturer"),
        ("Made in",                 "made_in"),
        ("Handling instruction",    "handling_instruction"),
    ]
    for label, key in add_pairs:
        _w(ws, r, 1, label, font=_font(bold=True))
        _m(ws, r, 2, r, 5, val=dn.get(key) or "", font=_font())
        r += 1

    r += 1
    # Signature
    _m(ws, r, 1, r, 3, val="FOR GODAVARI BIOREFINERIES B.V.", font=_font(bold=True))
    r += 2
    _w(ws, r, 1, "Authorised Signatory", font=_font())
    r += 2
    # Footer
    _m(ws, r, 1, r, 5,
       val="Commercial Register Number 34325188  |  VAT : NL8203.86.157.B.01",
       font=_font(size=8, color="555555"), align=_align("center"))
    r += 1
    _m(ws, r, 1, r, 5,
       val="Godavari Biorefineries B.V.  ·  Opaallaan 1180, 2132 LN Hoofddorp, The Netherlands  ·  Tel: +31 6 11 12 61 66",
       font=_font(size=8, color="555555"), align=_align("center"))

    buf = io.BytesIO()
    wb.save(buf)
    fname = f"DeliveryNote_{dn.get('reference_number') or 'draft'}.xlsx"
    return buf.getvalue(), fname
