"""CSV and Excel export for summary and shipping sheets."""
import csv
import io
from typing import Any, Optional

from openpyxl import Workbook

from app.database import format_quantity_display
from app.xl_style import (
    auto_col_widths,
    freeze_header,
    style_data_row,
    style_header_row,
)

ROLLUP_PRODUCT_COLUMNS = [
    ("Product", "product"),
    ("Customers", "customers"),
    ("Activities", "activities"),
    ("Last activity", "last_activity"),
]

ROLLUP_CUSTOMER_COLUMNS = [
    ("Customer", "customer"),
    ("Products", "products"),
    ("Activities", "activities"),
    ("Last activity", "last_activity"),
]

SHIPPING_COLUMNS = [
    ("Company", "company"),
    ("Product", "product"),
    ("Status", "status"),
    ("PO #", "po_number"),
    ("PO date", "po_date"),
    ("Quantity", "_quantity_display"),
    ("Packing", "packing"),
    ("GBL invoice", "gbl_invoice"),
    ("GBL invoice date", "gbl_invoice_date"),
    ("Container #", "container_number"),
    ("Vessel", "vessel_name"),
    ("ETD India", "etd_india"),
    ("Transit time", "transit_time"),
    ("Destination", "destination"),
    ("ETA", "eta"),
    ("Deal ID", "deal_id"),
]


def _cell(row: dict[str, Any], key: str) -> Any:
    if key == "_quantity_display":
        return format_quantity_display(
            row.get("quantity") or "",
            row.get("quantity_unit") or "MT",
        )
    val = row.get(key)
    return "" if val is None else val


def project_rows(
    rows: list[dict[str, Any]],
    columns: list[tuple[str, str]],
) -> list[dict[str, Any]]:
    return [{label: _cell(row, key) for label, key in columns} for row in rows]


def rollup_columns(group: str) -> list[tuple[str, str]]:
    return ROLLUP_CUSTOMER_COLUMNS if group == "customer" else ROLLUP_PRODUCT_COLUMNS


def rollup_sheet_name(group: str) -> str:
    return "By customer" if group == "customer" else "By product"


def to_csv_bytes(rows: list[dict[str, Any]], columns: list[tuple[str, str]]) -> bytes:
    projected = project_rows(rows, columns)
    buf = io.StringIO()
    if not projected:
        writer = csv.writer(buf)
        writer.writerow([label for label, _ in columns])
    else:
        writer = csv.DictWriter(buf, fieldnames=[label for label, _ in columns])
        writer.writeheader()
        writer.writerows(projected)
    return buf.getvalue().encode("utf-8-sig")


def to_xlsx_bytes(
    sheets: list[tuple[str, list[dict[str, Any]], list[tuple[str, str]]]],
) -> bytes:
    wb = Workbook()
    # Remove default empty sheet.
    wb.remove(wb.active)

    for sheet_name, rows, columns in sheets:
        projected = project_rows(rows, columns)
        safe_name = sheet_name[:31]
        ws = wb.create_sheet(title=safe_name)

        headers = [label for label, _ in columns]
        ws.append(headers)

        for row in projected:
            ws.append([row.get(h, "") for h in headers])

        n_cols = len(columns)
        style_header_row(ws, row=1, col_start=1, col_end=n_cols)
        ws.row_dimensions[1].height = 18

        for r_idx in range(2, ws.max_row + 1):
            style_data_row(ws, row=r_idx, col_start=1, col_end=n_cols,
                           alternate=(r_idx % 2 == 0))
            ws.row_dimensions[r_idx].height = 15

        auto_col_widths(ws, col_start=1, col_end=n_cols)
        freeze_header(ws, row=1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def export_filename(
    base: str,
    period: str,
    ext: str,
    group: Optional[str] = None,
) -> str:
    parts = [base]
    if group:
        parts.append(group)
    if period and period != "all":
        parts.append(period)
    return "-".join(parts) + f".{ext}"
