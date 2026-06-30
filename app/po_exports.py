"""Purchase Order Excel and PDF export."""
from __future__ import annotations

import shutil
from io import BytesIO
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.drawing.image import Image as XLImage

from app.database import get_data_dir
from app.purchase_orders import _float, safe_document_filename

BASE = Path(__file__).resolve().parent.parent
XLSX_DIR = get_data_dir() / "exports" / "purchase_orders" / "xlsx"
PDF_DIR  = get_data_dir() / "exports" / "purchase_orders" / "pdf"
# Template: prefer bundled seed dir, fall back to source tree.
_seed = __import__("os").environ.get("LEADS_SEED_DIR") or str(BASE / "data")
TEMPLATE_DIR = Path(_seed) / "templates" / "gbbv"
TEMPLATE_SRC = Path("/Users/padmajaganapathy/Documents/Godavari/Leads summary/GBBV/PO sample gbbv.xlsx")
TEMPLATE_FILE = TEMPLATE_DIR / "PO sample gbbv.xlsx"
LOGO_SVG = BASE / "static" / "gbbv-logo.svg"
LOGO_PNG = BASE / "static" / "gbbv-logo.png"


def _ensure_template() -> Path:
    TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)
    if not TEMPLATE_FILE.exists():
        if TEMPLATE_SRC.exists():
            shutil.copy2(TEMPLATE_SRC, TEMPLATE_FILE)
        else:
            raise FileNotFoundError(f"PO template not found: {TEMPLATE_SRC}")
    return TEMPLATE_FILE


def _batch_label(batch: dict[str, Any]) -> str:
    name = (batch.get("batch_name") or "").strip()
    qty = _float(batch.get("batch_quantity"))
    unit = batch.get("batch_unit") or "MT"
    if not name and not qty:
        return ""
    if qty:
        return f"{name} {qty:g} {unit}".strip()
    return name


def _fill_po_sheet(ws, po: dict[str, Any]) -> None:
    line = (po.get("line_items") or [{}])[0]
    batches = line.get("batches") or []

    ws["C2"] = po.get("company_name") or "Godavari Biorefineries Inc"
    ws["A5"] = po.get("document_title") or "Purchase Order"
    ws["A7"] = po.get("issuer_name")
    ws["E7"] = po.get("po_number")
    ws["A8"] = po.get("address_line_1")
    ws["E8"] = po.get("po_date")
    ws["A9"] = po.get("address_line_2")
    ws["E9"] = po.get("additional_ref")
    ws["A10"] = po.get("phone_1")
    ws["E10"] = po.get("payment_terms")
    ws["A11"] = po.get("phone_2")
    ws["E11"] = po.get("port_of_discharge")
    ws["A12"] = po.get("contact_person")
    ws["E12"] = po.get("incoterm_terms")
    ws["A13"] = po.get("email")
    ws["E13"] = po.get("shipment_timing")

    ws["A17"] = line.get("product_description")
    ws["B17"] = line.get("quantity_display")
    ws["C17"] = _float(line.get("pricing_quantity"))
    ws["D17"] = _float(line.get("rate"))
    ws["E17"] = _float(line.get("calculated_value"))
    ws["F17"] = line.get("remark")

    if len(batches) > 0:
        ws["F18"] = _batch_label(batches[0])
    ws["B19"] = "pack size : "
    ws["C19"] = _float(line.get("pack_size"))
    if len(batches) > 1:
        ws["F19"] = _batch_label(batches[1])
    ws["B20"] = "Number of pack"
    ws["C20"] = _float(line.get("number_of_packs"))
    ws["D20"] = line.get("incoterm_delivery_term") or po.get("incoterm_terms")
    ws["E20"] = _float(line.get("calculated_value"))
    if len(batches) > 2:
        label = _batch_label(batches[2])
        if len(batches) > 3:
            label += " | " + " | ".join(_batch_label(b) for b in batches[3:])
        ws["F20"] = label

    ws["A21"] = f"Express BL : Consignee: {po.get('consignee_name') or ''}"
    ws["A22"] = (
        f"{po.get('consignee_address') or ''} : Contact person : "
        f"{po.get('consignee_contact') or ''} : ph: {po.get('consignee_phone') or ''}"
    )
    ws["A24"] = (
        f"Notify : {po.get('notify_party') or ''} : Contact "
        f"{po.get('notify_contact') or ''}, {po.get('notify_address') or ''}"
    )
    ws["A25"] = f"HS code : {po.get('hs_code') or ''}"
    ws["A26"] = f"Credit note : remark : {po.get('credit_note_remark') or ''}"

    ws["A30"] = po.get("marking_buyer_name") or "Name of Buyer"
    ws["A31"] = po.get("marking_product_brand") or "Name of product + Brand Name"
    ws["A32"] = po.get("additional_ref") or po.get("po_number")
    ws["A33"] = po.get("marking_batch_no") or "Batch No"
    ws["A34"] = po.get("marking_gross_weight") or "Gross Weight"
    ws["A35"] = po.get("marking_net_weight") or "Net Weight"
    ws["A36"] = po.get("marking_tare_weight") or "Tare weight"
    ws["A37"] = po.get("marking_made_in") or "Made in India"
    ws["A38"] = po.get("marking_batch_on_docs") or "Batch number to be shown on all documents"
    ws["A39"] = po.get("marking_compliance_remark") or ""
    ws["A40"] = po.get("marking_loading_remark") or ""
    ws["A41"] = po.get("marking_inform_remark") or ""
    ws["A42"] = po.get("marking_pallets") or "Pallets : ISPM 3"

    docs = [d.strip() for d in (po.get("documents_required") or "").splitlines() if d.strip()]
    for i, doc in enumerate(docs):
        ws.cell(row=30 + i, column=3, value=doc)


def _add_logo(ws) -> None:
    if not LOGO_PNG.exists():
        return
    try:
        img = XLImage(str(LOGO_PNG))
        img.width = 80
        img.height = 60
        ws.add_image(img, "A1")
    except Exception:
        pass  # Pillow not available or image unreadable — skip silently


def build_po_workbook(po: dict[str, Any]) -> openpyxl.Workbook:
    template = _ensure_template()
    wb = openpyxl.load_workbook(template)
    ws = wb.active
    _fill_po_sheet(ws, po)
    _add_logo(ws)
    return wb


def export_po_xlsx(po: dict[str, Any]) -> tuple[bytes, str]:
    XLSX_DIR.mkdir(parents=True, exist_ok=True)
    fname = f"PO_{safe_document_filename(po.get('po_number', 'PO'))}.xlsx"
    path = XLSX_DIR / fname
    wb = build_po_workbook(po)
    wb.save(path)
    return path.read_bytes(), fname


def export_po_pdf_html(html: str) -> tuple[bytes, str] | None:
    from app.pdf_render import render_html_to_pdf

    pdf = render_html_to_pdf(html, base=BASE)
    if pdf:
        return pdf, "playwright"

    try:
        from weasyprint import HTML

        return HTML(string=html, base_url=str(BASE)).write_pdf(), "weasyprint"
    except Exception:
        pass
    try:
        from xhtml2pdf import pisa

        buf = BytesIO()
        pisa.CreatePDF(html, dest=buf, encoding="utf-8")
        return buf.getvalue(), "xhtml2pdf"
    except Exception:
        pass
    return None


def export_po_pdf(po: dict[str, Any], html: str) -> tuple[bytes, str] | None:
    result = export_po_pdf_html(html)
    if not result:
        return None
    pdf_bytes, _ = result
    PDF_DIR.mkdir(parents=True, exist_ok=True)
    fname = f"PO_{safe_document_filename(po.get('po_number', 'PO'))}.pdf"
    (PDF_DIR / fname).write_bytes(pdf_bytes)
    return pdf_bytes, fname
