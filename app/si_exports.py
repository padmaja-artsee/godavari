"""Commercial Invoice — Excel export.
Self-contained. Remove with the rest of the SI feature.
"""

from __future__ import annotations

import io
from typing import Any

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter


# ── style helpers ─────────────────────────────────────────────────────────────

_GREEN   = "1A5632"
_LT_GREEN = "E6F4EC"
_WHITE   = "FFFFFF"
_BORDER_CLR = "A0B8A8"

_thin = Side(style="thin", color=_BORDER_CLR)
_thick = Side(style="medium", color=_GREEN)

def _border(left=False, right=False, top=False, bottom=False, thick=False):
    side = _thick if thick else _thin
    return Border(
        left=side if left else Side(style=None),
        right=side if right else Side(style=None),
        top=side if top else Side(style=None),
        bottom=side if bottom else Side(style=None),
    )

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=10, color="000000", italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _write(ws, row, col, value, font=None, fill=None, align=None, border=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:   cell.font   = font
    if fill:   cell.fill   = fill
    if align:  cell.alignment = align
    if border: cell.border = border
    return cell

def _merge(ws, r1, c1, r2, c2, value=None, font=None, fill=None, align=None, border=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1, value=value)
    if font:   cell.font      = font
    if fill:   cell.fill      = fill
    if align:  cell.alignment = align
    if border: cell.border    = border
    return cell

def _num(v) -> float | str:
    try:
        f = float(str(v or "").replace(",", ""))
        return f if f != 0 else ""
    except (ValueError, TypeError):
        return v or ""


# ── main export ───────────────────────────────────────────────────────────────

def export_si_xlsx(si: dict[str, Any]) -> tuple[bytes, str]:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Commercial Invoice"

    # Column widths: A-G
    col_widths = [32, 12, 16, 16, 20, 0, 0]
    for i, w in enumerate(col_widths, 1):
        if w:
            ws.column_dimensions[get_column_letter(i)].width = w

    r = 1   # current row counter

    # ── Row 1: Company name ─────────────────────────────────────────
    _merge(ws, r, 1, r, 5,
           value=si.get("company_name") or "Godavari Biorefineries B.V.",
           font=_font(bold=True, size=13, color=_GREEN),
           align=_align("right"),
           fill=_fill(_WHITE))
    r += 1

    # ── Row 2: Document title ────────────────────────────────────────
    _merge(ws, r, 1, r, 5,
           value=si.get("document_title") or "COMMERCIAL INVOICE",
           font=_font(bold=True, size=14, color=_WHITE),
           fill=_fill(_GREEN),
           align=_align("center"))
    r += 1

    # ── Invoice No / Date ────────────────────────────────────────────
    _write(ws, r, 4, "Invoice No",  font=_font(bold=True))
    _write(ws, r, 5, si.get("invoice_number") or "",  font=_font())
    r += 1
    _write(ws, r, 4, "Date",        font=_font(bold=True))
    _write(ws, r, 5, si.get("invoice_date") or "",    font=_font())
    r += 1

    # ── Bill To ──────────────────────────────────────────────────────
    _write(ws, r, 1, "TO", font=_font(bold=True, size=11))
    r += 1
    for field in ("bill_to_name", "bill_to_address_1", "bill_to_address_2", "bill_to_address_3"):
        v = si.get(field) or ""
        if v:
            _merge(ws, r, 1, r, 3, value=v, font=_font(bold=(field == "bill_to_name")))
            r += 1
    for label, key in [("VAT Number", "bill_to_vat"),
                        ("Customer Order No", "customer_order_no"),
                        ("Code", "customer_material_code")]:
        v = si.get(key) or ""
        _write(ws, r, 1, label, font=_font(bold=True))
        _merge(ws, r, 2, r, 3, value=v, font=_font())
        r += 1

    # ── Transaction / Delivery ────────────────────────────────────────
    r += 1
    _write(ws, r, 1, "TRANSACTION DESCRIPTION", font=_font(bold=True))
    _merge(ws, r, 2, r, 5, value=si.get("transaction_description") or "", font=_font())
    r += 1
    for label, key in [("Delivery Address", "delivery_address"),
                        ("VAT of delivery location", "delivery_vat"),
                        ("Date of delivery", "delivery_date"),
                        ("ISO tank number", "iso_tank_number"),
                        ("Cleaning certificate no", "cleaning_cert_number")]:
        v = si.get(key) or ""
        _write(ws, r, 1, label, font=_font(bold=True))
        _merge(ws, r, 2, r, 5, value=v, font=_font())
        r += 1

    # ── Product table header ──────────────────────────────────────────
    r += 1
    hdr_fill = _fill(_GREEN)
    hdr_font = _font(bold=True, color=_WHITE)
    hdr_align = _align("center")
    qty_unit  = si.get("qty_unit") or "MT"
    r_cur     = si.get("rate_currency") or "Euro"
    r_unit    = si.get("rate_unit") or "MT"
    v_cur     = si.get("value_currency") or "Euro"

    col_headers = [
        ("PRODUCT AND OTHER DESCRIPTION", 1),
        (f"QUANTITY\n{qty_unit}", 2),
        (f"RATE\n{r_cur} per {r_unit}", 3),
        (f"VALUE\n{v_cur}", 4),
        ("TERMS / REMARK", 5),
    ]
    for label, c in col_headers:
        cell = ws.cell(row=r, column=c, value=label)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = _align("center", wrap=True)
        cell.border    = _border(True, True, True, True)
    ws.row_dimensions[r].height = 32
    r += 1

    # Brand name row
    if si.get("brand_name") or si.get("incoterms"):
        _write(ws, r, 1, f"Brand: {si.get('brand_name') or ''}", font=_font(italic=True))
        _write(ws, r, 5, si.get("incoterms") or "", font=_font())
        r += 1

    # Line items
    row_fill = _fill(_LT_GREEN)
    for i, line in enumerate(si.get("line_items") or []):
        f = row_fill if i % 2 == 0 else _fill(_WHITE)
        _write(ws, r, 1, line.get("product_description") or "", font=_font(), fill=f)
        _write(ws, r, 2, _num(line.get("quantity")),  font=_font(), fill=f, align=_align("right"))
        _write(ws, r, 3, _num(line.get("rate")),      font=_font(), fill=f, align=_align("right"))
        _write(ws, r, 4, _num(line.get("value")),     font=_font(), fill=f, align=_align("right"))
        _write(ws, r, 5, line.get("remark") or "",    font=_font(), fill=f)
        r += 1

    # Delivery note / batch rows
    if si.get("delivery_note_number"):
        _merge(ws, r, 1, r, 2,
               value=f"Delivery Note {si.get('delivery_note_number')}",
               font=_font(italic=True))
        r += 1
    if si.get("batch_numbers"):
        _write(ws, r, 1, f"Batch: {si.get('batch_numbers')}", font=_font(italic=True))
        _write(ws, r, 3, "Total", font=_font(bold=True), align=_align("right"))
        r += 1

    # ── Totals ────────────────────────────────────────────────────────
    r += 1
    tot_font = _font(bold=True)
    for label, val in [
        ("Net VALUE",    si.get("net_value") or 0),
        (f"VAT ({int(si.get('vat_percent') or 0)} %)", si.get("vat_amount") or 0),
        ("TOTAL TO PAY", si.get("total_to_pay") or 0),
    ]:
        _write(ws, r, 1, label, font=tot_font)
        _write(ws, r, 4, val,   font=tot_font, align=_align("right"))
        r += 1

    _write(ws, r, 1, "Total Payable In Words", font=tot_font)
    _merge(ws, r, 2, r, 5, value=si.get("amount_in_words") or "", font=_font())
    r += 1

    # ── Terms & Payment ────────────────────────────────────────────────
    r += 1
    for label, key in [("Terms of Delivery", "terms_of_delivery"),
                        ("Payment Terms",    "payment_terms")]:
        _write(ws, r, 1, label, font=_font(bold=True))
        _merge(ws, r, 2, r, 5, value=si.get(key) or "", font=_font())
        r += 1

    _write(ws, r, 1, "Enclosures", font=_font(bold=True))
    _merge(ws, r, 2, r, 5, value=si.get("enclosures") or "", font=_font(), align=_align(wrap=True))
    r += 1

    # ── Bank Details ───────────────────────────────────────────────────
    r += 1
    _merge(ws, r, 1, r, 5, value="BANK DETAILS",
           font=_font(bold=True, color=_WHITE), fill=_fill(_GREEN), align=_align("center"))
    r += 1
    for label, key in [("Bank",       "bank_name"),
                        ("Account No", "bank_account_no"),
                        ("IBAN",       "bank_iban"),
                        ("BIC",        "bank_bic")]:
        _write(ws, r, 1, label, font=_font(bold=True))
        _merge(ws, r, 2, r, 5, value=si.get(key) or "", font=_font())
        r += 1

    # ── Signature ─────────────────────────────────────────────────────
    r += 1
    _merge(ws, r, 1, r, 3, value="For Godavari Biorefineries B.V.", font=_font(bold=True))
    r += 2
    _write(ws, r, 1, "Authorised Signatory", font=_font())
    r += 1

    # ── Footer ────────────────────────────────────────────────────────
    r += 1
    _merge(ws, r, 1, r, 5,
           value="Commercial Register Number 34325188  |  VAT : NL8203.86.157.B.01",
           font=_font(size=8, color="555555"), align=_align("center"))
    r += 1
    _merge(ws, r, 1, r, 5,
           value="Godavari Biorefineries B.V.  ·  Opaallaan 1180, 2132 LN Hoofddorp, The Netherlands  ·  Tel: +31 6 11 12 61 66",
           font=_font(size=8, color="555555"), align=_align("center"))

    # ── Finalize ──────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    fname = f"CommercialInvoice_{si.get('invoice_number') or 'draft'}.xlsx"
    return buf.getvalue(), fname
