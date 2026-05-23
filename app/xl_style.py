"""Shared openpyxl styling utilities for all Excel exports."""
from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Palette ────────────────────────────────────────────────────────────────
BRAND_GREEN   = "1A5632"   # dark green header text / accents
HEADER_BG     = "1A5632"   # header row background
HEADER_FG     = "FFFFFF"   # header row text
SUBHEAD_BG    = "E8F5EE"   # sub-header / section title background
SUBHEAD_FG    = "1A5632"
ALT_ROW_BG    = "F5FAF7"   # alternating data row tint
BORDER_COLOR  = "B0C4B8"   # thin border colour

# ── Reusable style objects ─────────────────────────────────────────────────
_thin  = Side(style="thin",   color=BORDER_COLOR)
_thick = Side(style="medium", color="888888")

BORDER      = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
BORDER_THICK = Border(left=_thick, right=_thick, top=_thick, bottom=_thick)

HEADER_FONT   = Font(bold=True, color=HEADER_FG, size=10)
SUBHEAD_FONT  = Font(bold=True, color=SUBHEAD_FG, size=9)
DATA_FONT     = Font(size=9)
BOLD_FONT     = Font(bold=True, size=9)

HEADER_FILL  = PatternFill("solid", fgColor=HEADER_BG)
SUBHEAD_FILL = PatternFill("solid", fgColor=SUBHEAD_BG)
ALT_FILL     = PatternFill("solid", fgColor=ALT_ROW_BG)
WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right",  vertical="center", wrap_text=True)


# ── Helper functions ───────────────────────────────────────────────────────

def style_header_row(ws, row: int, col_start: int, col_end: int) -> None:
    """Apply green header styling to a full row range."""
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.border    = BORDER
        cell.alignment = CENTER


def style_subhead_row(ws, row: int, col_start: int, col_end: int) -> None:
    """Light green sub-header (section labels)."""
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.font      = SUBHEAD_FONT
        cell.fill      = SUBHEAD_FILL
        cell.border    = BORDER
        cell.alignment = CENTER


def style_data_row(ws, row: int, col_start: int, col_end: int,
                   alternate: bool = False) -> None:
    """Style a data row; alternate rows get a faint tint."""
    fill = ALT_FILL if alternate else WHITE_FILL
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.font   = DATA_FONT
        cell.fill   = fill
        cell.border = BORDER
        if cell.alignment.horizontal not in ("right", "center"):
            cell.alignment = LEFT


def auto_col_widths(ws, col_start: int, col_end: int,
                    min_w: int = 8, max_w: int = 42) -> None:
    """Set column widths based on content."""
    for c in range(col_start, col_end + 1):
        col_letter = get_column_letter(c)
        best = min_w
        for row_cells in ws.iter_rows(min_col=c, max_col=c):
            for cell in row_cells:
                if cell.value:
                    best = max(best, min(len(str(cell.value)) + 2, max_w))
        ws.column_dimensions[col_letter].width = best


def freeze_header(ws, row: int = 1) -> None:
    ws.freeze_panes = ws.cell(row=row + 1, column=1)
