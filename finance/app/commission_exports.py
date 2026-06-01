"""Monthly GBINC commission schedule Excel — matches finance/assets/commission_template.xlsx."""
from __future__ import annotations

import calendar
import os
import re
from copy import copy
from io import BytesIO
from pathlib import Path
from typing import Any

import openpyxl

_bundle = os.environ.get("LEADS_BUNDLE_BASE")
if _bundle:
    TEMPLATE = Path(_bundle) / "finance" / "assets" / "commission_template.xlsx"
else:
    TEMPLATE = Path(__file__).resolve().parent.parent / "assets" / "commission_template.xlsx"

_MONTH_UPPER = {m: calendar.month_name[m].upper() for m in range(1, 13)}


def _parse_amount(text: str) -> float:
    if not text:
        return 0.0
    s = str(text).replace(",", "").strip()
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    return float(m.group()) if m else 0.0


def _sail_date_display(deal: dict) -> str:
    raw = (
        deal.get("shipped_date")
        or deal.get("gbl_invoice_date")
        or deal.get("deal_date")
        or ""
    )[:10]
    if not raw or len(raw) < 10:
        return ""
    y, m, d = raw.split("-")
    return f"{d}.{m}.{y}"


def _sail_month_label(deal: dict, month: int) -> str:
    if month and month in _MONTH_UPPER:
        return _MONTH_UPPER[month]
    raw = (
        deal.get("shipped_date")
        or deal.get("gbl_invoice_date")
        or deal.get("deal_date")
        or ""
    )[:10]
    if len(raw) >= 7:
        try:
            m = int(raw[5:7])
            return _MONTH_UPPER.get(m, raw[5:7])
        except ValueError:
            pass
    return ""


def _invoice_number(deal: dict) -> str:
    return (deal.get("gbl_invoice") or "").strip()


def _deal_row_values(deal: dict, sr_no: int, month: int) -> dict[str, Any]:
    qty = _parse_amount(deal.get("quantity") or "")
    rate = _parse_amount(deal.get("price") or "")
    if not rate and deal.get("value"):
        # Fallback: cannot split value without qty
        pass
    currency = "USD"
    price_raw = (deal.get("price") or "") + (deal.get("price_unit") or "")
    if "EUR" in price_raw.upper():
        currency = "EUR"
    return {
        "sr_no": sr_no,
        "sail_month": _sail_month_label(deal, month),
        "invoice_no": _invoice_number(deal),
        "vessel_sail_date": _sail_date_display(deal),
        "ship_to": deal.get("company") or "",
        "port": (deal.get("destination") or "").strip() or "—",
        "qty": qty,
        "currency": currency,
        "rate": rate,
    }


def export_commission_xlsx(
    deals: list[dict],
    *,
    product_label: str,
    period_label: str,
    month: int,
) -> tuple[bytes, str]:
    if not TEMPLATE.exists():
        raise FileNotFoundError(f"Commission template missing: {TEMPLATE}")

    wb = openpyxl.load_workbook(TEMPLATE)
    ws = wb.active

    prod = (product_label or "PRODUCTS").strip().upper()
    ws["B2"] = f"GBINC COMMISSION TOWARDS SUPPLY OF {prod}"

    first_data_row = 4
    n = len(deals)
    if n > 1:
        ws.insert_rows(5, n - 1)  # template row 5 is TOTAL
    total_row = first_data_row + max(n, 1)

    sample_style = {}
    if n:
        sample_style = {c: copy(ws.cell(first_data_row, c)) for c in range(2, 16)}

    for i, deal in enumerate(deals):
        r = first_data_row + i
        vals = _deal_row_values(deal, i + 1, month)
        mapping = [
            (2, vals["sr_no"]),
            (3, vals["sail_month"]),
            (4, vals["invoice_no"]),
            (5, vals["vessel_sail_date"]),
            (6, vals["ship_to"]),
            (7, vals["port"]),
            (8, vals["qty"]),
            (9, vals["currency"]),
            (10, vals["rate"]),
        ]
        for col, val in mapping:
            cell = ws.cell(r, col, val)
            if col in sample_style:
                ref = sample_style[col]
                cell.font = copy(ref.font)
                cell.fill = copy(ref.fill)
                cell.border = copy(ref.border)
                cell.alignment = copy(ref.alignment)
                cell.number_format = ref.number_format
        # VALUE = QTY × RATE; FOB = VALUE − freight − insurance; COMMISSION = FOB × 3%
        ws.cell(r, 11).value = f"=H{r}*J{r}"
        ws.cell(r, 11).number_format = "#,##0.00"
        ws.cell(r, 14).value = f"=K{r}-L{r}-M{r}"
        ws.cell(r, 14).number_format = "#,##0.00"
        ws.cell(r, 15).value = f"=N{r}*3%"
        ws.cell(r, 15).number_format = "#,##0.00"

    if n:
        ws.cell(total_row, 2, "TOTAL")
        ws.cell(total_row, 2).font = openpyxl.styles.Font(bold=True)
        ws.merge_cells(start_row=total_row, start_column=2, end_row=total_row, end_column=7)
        for col, letter in ((8, "H"), (11, "K"), (14, "N"), (15, "O")):
            ws.cell(total_row, col).value = f"=SUM({letter}{first_data_row}:{letter}{total_row - 1})"
            ws.cell(total_row, col).font = openpyxl.styles.Font(bold=True)
            ws.cell(total_row, col).number_format = "#,##0.00"
    else:
        ws.cell(first_data_row, 2, "— no deals for this period —")

    safe_period = period_label.replace(" ", "_").replace("/", "-")[:40]
    fname = f"GBINC_Commission_{safe_period}.xlsx"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue(), fname
