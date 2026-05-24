"""Excel export — budget template, grid exports, list exports."""
from __future__ import annotations
from io import BytesIO
from typing import List

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from finance.app.database import (
    FY_MONTHS, MONTH_LABELS, SECTION_LABELS, SECTION_ORDER,
    compute_grid, get_budget_grid, get_transaction_rollup,
    get_actuals_manual,
)

_GREEN   = "1C5631"
_GREEN_D = "134023"
_GREEN_L = "D9EAD3"
_GREY    = "F2F2F2"
_WHITE   = "FFFFFF"
_BLACK   = "000000"
_RED_L   = "FCE4E4"
_RED     = "C0392B"
_GREEN_C = "E8F5E9"


def _f(bold=False, size=9, color=_BLACK, name="Calibri"):
    return Font(name=name, size=size, bold=bold, color=color)

def _fill(c):
    return PatternFill("solid", fgColor=c)

def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _al(h="right", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _write_section_header(ws, row, label, ncols):
    ws.merge_cells(f"B{row}:{get_column_letter(ncols)}{row}")
    c = ws.cell(row=row, column=2, value=label)
    c.font = _f(bold=True, size=9, color=_WHITE)
    c.fill = _fill(_GREEN_D)
    c.alignment = _al(h="left")
    ws.row_dimensions[row].height = 14


def generate_budget_template(items: list[dict], fiscal_year: int) -> tuple[bytes, str]:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Budget FY{str(fiscal_year)[-2:]}"

    ncols = 15  # B=label, C-N=months, O=Total
    month_col = {m: 3 + i for i, m in enumerate(FY_MONTHS)}

    # Title
    ws.merge_cells(f"B1:{get_column_letter(ncols)}1")
    c = ws["B1"]
    c.value = f"Godavari Biorefineries Inc"
    c.font = _f(bold=True, size=12, color=_WHITE)
    c.fill = _fill(_GREEN)
    c.alignment = _al(h="left")
    ws.row_dimensions[1].height = 22

    ws.merge_cells(f"B2:{get_column_letter(ncols)}2")
    c = ws["B2"]
    c.value = f"PLANNED EXPENSES — FY{fiscal_year} (Apr {fiscal_year-1}–Mar {fiscal_year})"
    c.font = _f(bold=True, size=10, color=_WHITE)
    c.fill = _fill(_GREEN)
    c.alignment = _al(h="left")
    ws.row_dimensions[2].height = 18

    # Column headers
    hdr_row = 4
    ws.cell(row=hdr_row, column=2, value="").fill = _fill(_GREEN)
    for m in FY_MONTHS:
        c = ws.cell(row=hdr_row, column=month_col[m],
                    value=MONTH_LABELS[m].upper())
        c.font = _f(bold=True, size=8, color=_WHITE)
        c.fill = _fill(_GREEN)
        c.alignment = _al(h="center")
        c.border = _border()
    c = ws.cell(row=hdr_row, column=ncols, value="TOTAL")
    c.font = _f(bold=True, size=8, color=_WHITE)
    c.fill = _fill(_GREEN_D)
    c.alignment = _al(h="center")
    c.border = _border()
    ws.row_dimensions[hdr_row].height = 16

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 30
    for col in range(3, ncols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 10

    # Data rows
    data_row = hdr_row + 1
    prev_sec = None
    for item in items:
        sec = item["section"]
        if sec in ("income", "totals"):
            continue  # template is expense-only (matches original)
        is_calc = item["is_calculated"]
        if sec != prev_sec:
            _write_section_header(ws, data_row, SECTION_LABELS.get(sec, sec).upper(), ncols)
            data_row += 1
            # Sub-header
            ws.cell(row=data_row, column=2, value=SECTION_LABELS.get(sec, sec).upper()).font = _f(bold=True, size=8)
            for m in FY_MONTHS:
                c = ws.cell(row=data_row, column=month_col[m], value=MONTH_LABELS[m].upper())
                c.font = _f(bold=True, size=8)
                c.alignment = _al(h="center")
            ws.cell(row=data_row, column=ncols, value="TOTAL").font = _f(bold=True, size=8)
            data_row += 1
            prev_sec = sec

        row_fill = _fill(_GREEN_L) if is_calc else (_fill(_GREY) if data_row % 2 == 0 else _fill(_WHITE))
        row_font = _f(bold=is_calc)

        c = ws.cell(row=data_row, column=2, value=item["name"])
        c.font = row_font; c.fill = row_fill
        c.alignment = _al(h="left"); c.border = _border()

        for m in FY_MONTHS:
            c = ws.cell(row=data_row, column=month_col[m])
            c.fill = row_fill; c.border = _border()
            if is_calc:
                c.value = "(auto)"; c.font = _f(size=8, color="999999")
            else:
                c.number_format = "#,##0.00"; c.font = row_font
            c.alignment = _al(h="right")

        c = ws.cell(row=data_row, column=ncols)
        c.fill = _fill(_GREEN_L) if is_calc else _fill(_GREY)
        c.font = _f(bold=True)
        if not is_calc: c.value = 0; c.number_format = "#,##0.00"
        c.alignment = _al(h="right"); c.border = _border()
        ws.row_dimensions[data_row].height = 14
        data_row += 1

    # Monthly Total row
    data_row += 1
    ws.merge_cells(f"B{data_row}:{get_column_letter(ncols)}{data_row}")
    pass  # left blank for user totals

    ws.freeze_panes = "C5"
    fname = f"GBInc-Budget-Template-FY{fiscal_year}.xlsx"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue(), fname


def parse_budget_upload(file_bytes: bytes, items: list[dict]) -> dict:
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    name_to_item = {i["name"].strip().lower(): i for i in items if not i["is_calculated"]}
    values: dict = {}

    # Find header row
    header_row = None
    for r in ws.iter_rows():
        vals = [str(c.value or "").strip().lower() for c in r]
        if any(v in ("jan", "january", "apr", "april") for v in vals):
            header_row = r
            break
    if not header_row:
        raise ValueError("Could not find month header row.")

    label_to_month = {v.lower()[:3]: k for k, v in MONTH_LABELS.items()}
    label_to_month.update({v.lower(): k for k, v in MONTH_LABELS.items()})
    # also map full month names
    full = {"january":1,"february":2,"march":3,"april":4,"may":5,"june":6,
            "july":7,"august":8,"september":9,"october":10,"november":11,"december":12}
    label_to_month.update(full)

    col_to_month = {}
    for cell in header_row:
        lbl = str(cell.value or "").strip().lower()
        if lbl in label_to_month:
            col_to_month[cell.column] = label_to_month[lbl]

    if not col_to_month:
        raise ValueError("No month columns found.")

    for row in ws.iter_rows(min_row=header_row[0].row + 1):
        name = str(row[0].value or "").strip().lower()
        if not name:
            name = str(row[1].value or "").strip().lower() if len(row) > 1 else ""
        item = name_to_item.get(name)
        if not item:
            continue
        lid = item["id"]
        for cell in row:
            month = col_to_month.get(cell.column)
            if month is None:
                continue
            try:
                val = float(cell.value) if cell.value not in (None, "", "(auto)") else 0.0
            except (TypeError, ValueError):
                val = 0.0
            if val:
                values[f"{lid}_{month}"] = val
    return values


# ─────────────────────────────────────────────────────────────────────────────
# Shared helpers for grid exports
# ─────────────────────────────────────────────────────────────────────────────

def _grid_header(ws, title: str, subtitle: str, ncols: int) -> None:
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    c = ws["A1"]; c.value = "Godavari Biorefineries Inc"
    c.font = _f(bold=True, size=12, color=_WHITE); c.fill = _fill(_GREEN)
    c.alignment = _al(h="left"); ws.row_dimensions[1].height = 20

    ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
    c = ws["A2"]; c.value = title
    c.font = _f(bold=True, size=10, color=_WHITE); c.fill = _fill(_GREEN)
    c.alignment = _al(h="left"); ws.row_dimensions[2].height = 16

    ws.merge_cells(f"A3:{get_column_letter(ncols)}3")
    c = ws["A3"]; c.value = subtitle
    c.font = _f(size=9, color="666666"); c.alignment = _al(h="left")
    ws.row_dimensions[3].height = 14


SECTION_STYLES_XL = {
    "Total Income":            ("subtotal", _GREEN_L, True),
    "Total Employee Costs":    ("subtotal", _GREEN_L, True),
    "Total Office Costs":      ("subtotal", _GREEN_L, True),
    "Total Admin":             ("subtotal", _GREEN_L, True),
    "Total Travel":            ("subtotal", _GREEN_L, True),
    "Total Expenses":          ("highlight","FFF9C4",  True),
    "Net (Income - Expenses)": ("total",    _GREEN_L,  True),
}


def _write_grid_sheet(ws, items: list, grid: dict, title: str, subtitle: str) -> None:
    ncols = 2 + len(FY_MONTHS) + 1  # label + 12 months + total
    _grid_header(ws, title, subtitle, ncols)

    # Column headers row 5
    hdr = 5
    ws.cell(hdr, 1, "Line Item").font = _f(bold=True, size=8, color=_WHITE)
    ws.cell(hdr, 1).fill = _fill(_GREEN)
    for i, m in enumerate(FY_MONTHS, start=2):
        c = ws.cell(hdr, i, MONTH_LABELS[m].upper())
        c.font = _f(bold=True, size=8, color=_WHITE)
        c.fill = _fill(_GREEN); c.alignment = _al(h="center"); c.border = _border()
    c = ws.cell(hdr, ncols, "TOTAL")
    c.font = _f(bold=True, size=8, color=_WHITE)
    c.fill = _fill(_GREEN_D); c.alignment = _al(h="center"); c.border = _border()
    ws.row_dimensions[hdr].height = 14

    ws.column_dimensions["A"].width = 32
    for col in range(2, ncols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 10

    full = compute_grid(grid, items)
    data_row = hdr + 1
    prev_sec = None

    for item in items:
        sec = item["section"]
        if sec == "totals":
            data_row += 1  # spacer
        if sec != prev_sec and sec != "totals":
            # Section header
            ws.merge_cells(f"A{data_row}:{get_column_letter(ncols)}{data_row}")
            c = ws.cell(data_row, 1, SECTION_LABELS.get(sec, sec).upper())
            c.font = _f(bold=True, size=8, color=_WHITE); c.fill = _fill(_GREEN)
            ws.row_dimensions[data_row].height = 13
            data_row += 1; prev_sec = sec

        style_info = SECTION_STYLES_XL.get(item["name"])
        fill_color = style_info[1] if style_info else (_GREY if data_row % 2 == 0 else _WHITE)
        row_font = _f(bold=bool(style_info))

        c = ws.cell(data_row, 1, item["name"])
        c.font = row_font; c.fill = _fill(fill_color)
        c.alignment = _al(h="left"); c.border = _border()

        row_total = 0.0
        for i, m in enumerate(FY_MONTHS, start=2):
            val = full.get((item["id"], m), 0.0)
            row_total += val
            cell = ws.cell(data_row, i, val if val else None)
            cell.fill = _fill(fill_color); cell.border = _border()
            cell.number_format = "#,##0.00"; cell.alignment = _al(h="right")
            cell.font = row_font

        tot_cell = ws.cell(data_row, ncols, row_total if row_total else None)
        tot_cell.fill = _fill(_GREEN_L if style_info else _GREY)
        tot_cell.number_format = "#,##0.00"; tot_cell.alignment = _al(h="right")
        tot_cell.font = _f(bold=True); tot_cell.border = _border()
        ws.row_dimensions[data_row].height = 13
        data_row += 1

    ws.freeze_panes = "B6"


# ─────────────────────────────────────────────────────────────────────────────
# Budget export
# ─────────────────────────────────────────────────────────────────────────────

def export_budget_xlsx(items: list, fiscal_year: int) -> tuple:
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = f"Budget FY{fiscal_year}"
    grid = get_budget_grid(fiscal_year)
    _write_grid_sheet(ws, items, grid,
        f"Budget — FY{fiscal_year}",
        f"Apr {fiscal_year-1} – Mar {fiscal_year}")
    buf = BytesIO(); wb.save(buf)
    return buf.getvalue(), f"GBInc-Budget-FY{fiscal_year}.xlsx"


# ─────────────────────────────────────────────────────────────────────────────
# Actuals export
# ─────────────────────────────────────────────────────────────────────────────

def export_actuals_xlsx(items: list, fiscal_year: int) -> tuple:
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = f"Actuals FY{fiscal_year}"
    rollup  = get_transaction_rollup(fiscal_year)
    manual  = get_actuals_manual(fiscal_year)
    combined = {}
    for k in set(list(rollup.keys()) + list(manual.keys())):
        combined[k] = rollup.get(k, 0) + manual.get(k, 0)
    _write_grid_sheet(ws, items, combined,
        f"Actuals — FY{fiscal_year}",
        f"Apr {fiscal_year-1} – Mar {fiscal_year}")
    buf = BytesIO(); wb.save(buf)
    return buf.getvalue(), f"GBInc-Actuals-FY{fiscal_year}.xlsx"


# ─────────────────────────────────────────────────────────────────────────────
# Budget vs Actuals variances export
# ─────────────────────────────────────────────────────────────────────────────

def export_variances_xlsx(items: list, fiscal_year: int) -> tuple:
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = f"Variances FY{fiscal_year}"

    ncols = 2 + len(FY_MONTHS) * 3 + 3  # label + (bud/act/var)*12 + 3 totals
    _grid_header(ws, f"Budget vs Actuals — FY{fiscal_year}",
                 f"Apr {fiscal_year-1} – Mar {fiscal_year}", min(ncols, 50))

    b_grid = compute_grid(get_budget_grid(fiscal_year), items)
    rollup  = get_transaction_rollup(fiscal_year)
    manual  = get_actuals_manual(fiscal_year)
    combined = {}
    for k in set(list(rollup.keys()) + list(manual.keys())):
        combined[k] = rollup.get(k, 0) + manual.get(k, 0)
    a_grid = compute_grid(combined, items)

    # Sub-header rows
    hdr1, hdr2 = 5, 6
    ws.cell(hdr1, 1, "Line Item").font = _f(bold=True, size=8, color=_WHITE)
    ws.cell(hdr1, 1).fill = _fill(_GREEN)
    col = 2
    for m in FY_MONTHS:
        lbl = MONTH_LABELS[m].upper()
        ws.merge_cells(f"{get_column_letter(col)}{hdr1}:{get_column_letter(col+2)}{hdr1}")
        c = ws.cell(hdr1, col, lbl)
        c.font = _f(bold=True, size=7, color=_WHITE); c.fill = _fill(_GREEN)
        c.alignment = _al(h="center")
        for sub, nm in [(col,"Bud"),(col+1,"Act"),(col+2,"Var")]:
            c2 = ws.cell(hdr2, sub, nm)
            c2.font = _f(bold=True, size=7); c2.fill = _fill(_GREY)
            c2.alignment = _al(h="center"); c2.border = _border()
        col += 3
    for sub, nm in [(col,"Bud"),(col+1,"Act"),(col+2,"Var")]:
        c2 = ws.cell(hdr1, sub, "TOTAL" if sub==col else "")
        c3 = ws.cell(hdr2, sub, nm)
        c2.font = _f(bold=True, size=7, color=_WHITE); c2.fill = _fill(_GREEN_D)
        c3.font = _f(bold=True, size=7); c3.fill = _fill(_GREY)
        c3.alignment = _al(h="center"); c3.border = _border()

    ws.column_dimensions["A"].width = 30
    for c in range(2, col + 4):
        ws.column_dimensions[get_column_letter(c)].width = 8

    data_row = hdr2 + 1
    prev_sec = None
    for item in items:
        sec = item["section"]
        if sec == "totals": data_row += 1
        if sec != prev_sec and sec != "totals":
            ws.merge_cells(f"A{data_row}:{get_column_letter(col+2)}{data_row}")
            c = ws.cell(data_row, 1, SECTION_LABELS.get(sec, sec).upper())
            c.font = _f(bold=True, size=7, color=_WHITE); c.fill = _fill(_GREEN)
            data_row += 1; prev_sec = sec

        style_info = SECTION_STYLES_XL.get(item["name"])
        fill_color = style_info[1] if style_info else (_GREY if data_row % 2 == 0 else _WHITE)
        ws.cell(data_row, 1, item["name"]).font = _f(bold=bool(style_info))
        ws.cell(data_row, 1).fill = _fill(fill_color)
        ws.cell(data_row, 1).alignment = _al(h="left"); ws.cell(data_row, 1).border = _border()

        b_tot = a_tot = 0.0
        c = 2
        for m in FY_MONTHS:
            b = b_grid.get((item["id"], m), 0.0); a = a_grid.get((item["id"], m), 0.0)
            v = b - a if item["section"] != "income" else a - b
            b_tot += b; a_tot += a
            for ci, val in [(c,b),(c+1,a),(c+2,v)]:
                cell = ws.cell(data_row, ci, val if val else None)
                cell.number_format = "#,##0.00"; cell.border = _border()
                cell.fill = _fill(fill_color); cell.alignment = _al(h="right")
                if ci == c+2 and val:
                    cell.font = _f(bold=True, color=("006100" if val >= 0 else "9C0006"))
            c += 3
        v_tot = b_tot - a_tot if item["section"] != "income" else a_tot - b_tot
        for ci, val in [(c,b_tot),(c+1,a_tot),(c+2,v_tot)]:
            cell = ws.cell(data_row, ci, val if val else None)
            cell.number_format = "#,##0.00"; cell.border = _border()
            cell.fill = _fill(_GREEN_L if style_info else _GREY)
            cell.font = _f(bold=True, color=("006100" if ci==c+2 and v_tot>=0 else ("9C0006" if ci==c+2 else _BLACK)))
        data_row += 1

    ws.freeze_panes = "B7"
    buf = BytesIO(); wb.save(buf)
    return buf.getvalue(), f"GBInc-Variances-FY{fiscal_year}.xlsx"


# ─────────────────────────────────────────────────────────────────────────────
# Transactions list (expenses or income)
# ─────────────────────────────────────────────────────────────────────────────

def export_transactions_xlsx(transactions: list, fiscal_year: int,
                              tx_type: str = "expense") -> tuple:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Income" if tx_type == "income" else "Expenses"

    label = "Income Entries" if tx_type == "income" else "Expenses"
    _grid_header(ws, f"{label} — FY{fiscal_year}",
                 f"Apr {fiscal_year-1} – Mar {fiscal_year}", 8)

    headers = ["Date","Account","Vendor","Payment Account","Reference","Currency","Amount","Notes"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(5, i, h)
        c.font = _f(bold=True, size=8, color=_WHITE); c.fill = _fill(_GREEN)
        c.alignment = _al(h="center"); c.border = _border()

    widths = [12, 28, 22, 18, 16, 9, 12, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    total = 0.0
    for r, t in enumerate(transactions, 6):
        row_fill = _fill(_WHITE) if r % 2 == 0 else _fill(_GREY)
        vals = [
            t.get("date",""), t.get("account_name",""), t.get("vendor_name") or "",
            t.get("payment_account_name") or "", t.get("reference") or "",
            t.get("currency","USD"), t.get("amount",0), t.get("notes") or "",
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(r, ci, v)
            cell.fill = row_fill; cell.border = _border()
            if ci == 7:
                cell.number_format = "#,##0.00"; cell.alignment = _al(h="right")
            ws.row_dimensions[r].height = 13
        total += t.get("amount", 0)

    # Total row
    tr = 6 + len(transactions)
    ws.cell(tr, 6, "TOTAL").font = _f(bold=True)
    ws.cell(tr, 6).fill = _fill(_GREEN_L); ws.cell(tr, 6).border = _border()
    c = ws.cell(tr, 7, total)
    c.font = _f(bold=True); c.fill = _fill(_GREEN_L)
    c.number_format = "#,##0.00"; c.alignment = _al(h="right"); c.border = _border()

    ws.freeze_panes = "A6"
    fname = f"GBInc-{'Income' if tx_type=='income' else 'Expenses'}-FY{fiscal_year}.xlsx"
    buf = BytesIO(); wb.save(buf)
    return buf.getvalue(), fname


# ─────────────────────────────────────────────────────────────────────────────
# Vendors list
# ─────────────────────────────────────────────────────────────────────────────

def export_vendors_xlsx(vendors: list) -> tuple:
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Vendors"
    _grid_header(ws, "Vendors", "Godavari Biorefineries Inc", 5)

    for i, h in enumerate(["Name","Email","Phone","Notes","ID"], 1):
        c = ws.cell(5, i, h)
        c.font = _f(bold=True, size=8, color=_WHITE); c.fill = _fill(_GREEN)
        c.alignment = _al(h="center"); c.border = _border()

    for col, w in [(1,28),(2,26),(3,16),(4,36),(5,8)]:
        ws.column_dimensions[get_column_letter(col)].width = w

    for r, v in enumerate(vendors, 6):
        row_fill = _fill(_WHITE) if r % 2 == 0 else _fill(_GREY)
        for ci, val in enumerate([v.get("name",""), v.get("email",""),
                                    v.get("phone",""), v.get("notes",""),
                                    v.get("id","")], 1):
            cell = ws.cell(r, ci, val)
            cell.fill = row_fill; cell.border = _border()
            ws.row_dimensions[r].height = 13

    buf = BytesIO(); wb.save(buf)
    return buf.getvalue(), "GBInc-Vendors.xlsx"


# ─────────────────────────────────────────────────────────────────────────────
# P&L Report
# ─────────────────────────────────────────────────────────────────────────────

def export_report_xlsx(
    period_label: str,
    income_lines: list,
    total_income_actual: float,
    total_income_budget: float,
    expense_sections: list,
    total_expenses_actual: float,
    total_expenses_budget: float,
    net_actual: float,
    net_budget: float,
    fiscal_year: int,
) -> tuple:
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "P&L Report"

    # ── Header rows ───────────────────────────────────────────────────────────
    ws.merge_cells("A1:D1")
    c = ws["A1"]; c.value = "Godavari Biorefineries Inc"
    c.font = _f(bold=True, size=13, color=_WHITE); c.fill = _fill(_GREEN_D)
    c.alignment = _al(h="left"); ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:D2")
    c = ws["A2"]; c.value = f"Profit & Loss — {period_label}"
    c.font = _f(bold=True, size=11, color=_WHITE); c.fill = _fill(_GREEN)
    c.alignment = _al(h="left"); ws.row_dimensions[2].height = 18

    # ── Column headers ────────────────────────────────────────────────────────
    for ci, h in enumerate(["", "Actual", "Budget", "Variance"], 1):
        c = ws.cell(3, ci, h)
        c.font = _f(bold=True, size=9, color=_WHITE)
        c.fill = _fill(_GREEN); c.alignment = _al(h="right" if ci > 1 else "left")
        c.border = _border()
    ws.column_dimensions["A"].width = 32
    for col in ["B", "C", "D"]:
        ws.column_dimensions[col].width = 14
    ws.row_dimensions[3].height = 14

    def _money(ws, row, col, val, bold=False, fill=None):
        c = ws.cell(row, col, val if val else 0)
        c.number_format = "#,##0.00"
        c.alignment = _al(h="right")
        c.font = _f(bold=bold)
        if fill: c.fill = _fill(fill)
        c.border = _border()
        return c

    def _label(ws, row, val, indent=0, bold=False, fill=None):
        c = ws.cell(row, 1, ("  " * indent) + val)
        c.font = _f(bold=bold)
        if fill: c.fill = _fill(fill)
        c.border = _border()
        ws.row_dimensions[row].height = 13
        return c

    row = 4

    # ── INCOME ────────────────────────────────────────────────────────────────
    ws.cell(row, 1, "INCOME").font = _f(bold=True, size=9, color=_WHITE)
    ws.cell(row, 1).fill = _fill(_GREEN)
    ws.cell(row, 1).border = _border()
    for ci in range(2, 5):
        ws.cell(row, ci).fill = _fill(_GREEN); ws.cell(row, ci).border = _border()
    ws.row_dimensions[row].height = 14
    row += 1

    for line in income_lines:
        act, bud = line["actual"], line["budget"]
        _label(ws, row, line["name"], indent=1)
        _money(ws, row, 2, act); _money(ws, row, 3, bud)
        _money(ws, row, 4, act - bud)
        row += 1

    _label(ws, row, "Total Income", bold=True, fill=_GREEN_L)
    _money(ws, row, 2, total_income_actual, bold=True, fill=_GREEN_L)
    _money(ws, row, 3, total_income_budget, bold=True, fill=_GREEN_L)
    _money(ws, row, 4, total_income_actual - total_income_budget, bold=True, fill=_GREEN_L)
    row += 2  # blank spacer

    # ── EXPENSES ─────────────────────────────────────────────────────────────
    ws.cell(row, 1, "EXPENSES").font = _f(bold=True, size=9, color=_WHITE)
    ws.cell(row, 1).fill = _fill(_GREEN)
    ws.cell(row, 1).border = _border()
    for ci in range(2, 5):
        ws.cell(row, ci).fill = _fill(_GREEN); ws.cell(row, ci).border = _border()
    ws.row_dimensions[row].height = 14
    row += 1

    for sec in expense_sections:
        # section header
        _label(ws, row, sec["label"].upper(), bold=True, fill=_GREY)
        for ci in range(2, 5):
            ws.cell(row, ci).fill = _fill(_GREY); ws.cell(row, ci).border = _border()
        ws.row_dimensions[row].height = 13
        row += 1

        for line in sec["lines"]:
            act, bud = line["actual"], line["budget"]
            _label(ws, row, line["name"], indent=1)
            _money(ws, row, 2, act); _money(ws, row, 3, bud)
            _money(ws, row, 4, bud - act)
            row += 1

        # section total
        sa, sb = sec["total_actual"], sec["total_budget"]
        _label(ws, row, f"Total {sec['label']}", bold=True, fill=_GREEN_L)
        _money(ws, row, 2, sa, bold=True, fill=_GREEN_L)
        _money(ws, row, 3, sb, bold=True, fill=_GREEN_L)
        _money(ws, row, 4, sb - sa, bold=True, fill=_GREEN_L)
        row += 2  # spacer after section

    # Total Expenses
    _label(ws, row, "TOTAL EXPENSES", bold=True, fill="FFF9C4")
    _money(ws, row, 2, total_expenses_actual, bold=True, fill="FFF9C4")
    _money(ws, row, 3, total_expenses_budget, bold=True, fill="FFF9C4")
    _money(ws, row, 4, total_expenses_budget - total_expenses_actual, bold=True, fill="FFF9C4")
    row += 2

    # Net Income
    _label(ws, row, "NET INCOME (Income − Expenses)", bold=True, fill=_GREEN_L)
    _money(ws, row, 2, net_actual, bold=True, fill=_GREEN_L)
    _money(ws, row, 3, net_budget, bold=True, fill=_GREEN_L)
    _money(ws, row, 4, net_actual - net_budget, bold=True, fill=_GREEN_L)

    ws.freeze_panes = "B4"
    fname = f"GBInc-PnL-FY{fiscal_year}.xlsx"
    buf = BytesIO(); wb.save(buf)
    return buf.getvalue(), fname


# ─────────────────────────────────────────────────────────────────────────────
# Analysis (Budget vs Actuals summary + monthly breakdown)
# ─────────────────────────────────────────────────────────────────────────────

def export_analysis_xlsx(summary: list, chart_months: list,
                          exp_budget: list, exp_actual: list,
                          income_actual: list, fiscal_year: int) -> tuple:
    wb = openpyxl.Workbook()

    # ── Sheet 1: Section Summary ───────────────────────────────────────────
    ws1 = wb.active; ws1.title = "Summary"
    _grid_header(ws1, f"Budget vs Actuals Summary — FY{fiscal_year}",
                 "Godavari Biorefineries Inc", 5)

    for ci, h in enumerate(["Section", "Planned", "Actual", "Variance", "Variance %"], 1):
        c = ws1.cell(5, ci, h)
        c.font = _f(bold=True, size=8, color=_WHITE)
        c.fill = _fill(_GREEN); c.alignment = _al(h="center"); c.border = _border()
    for col, w in [(1, 22), (2, 14), (3, 14), (4, 14), (5, 12)]:
        ws1.column_dimensions[get_column_letter(col)].width = w

    for r, s in enumerate(summary, 6):
        rf = _fill(_WHITE) if r % 2 == 0 else _fill(_GREY)
        ws1.cell(r, 1, s["label"]).fill = rf; ws1.cell(r, 1).border = _border()
        for ci, key in enumerate(["planned", "actual", "variance"], 2):
            c = ws1.cell(r, ci, s.get(key, 0))
            c.number_format = "#,##0"; c.alignment = _al(h="right")
            c.fill = rf; c.border = _border()
        vp = s.get("var_pct")
        c5 = ws1.cell(r, 5, round(vp, 1) if vp is not None else "")
        if vp is not None:
            c5.number_format = '#,##0.0"%"'
        c5.alignment = _al(h="right"); c5.fill = rf; c5.border = _border()
        ws1.row_dimensions[r].height = 13
    ws1.freeze_panes = "A6"

    # ── Sheet 2: Monthly Expenses ──────────────────────────────────────────
    ws2 = wb.create_sheet("Monthly Expenses")
    _grid_header(ws2, f"Monthly Expenses — FY{fiscal_year}",
                 "Godavari Biorefineries Inc", 4)
    for ci, h in enumerate(["Month", "Budget", "Actuals", "Variance"], 1):
        c = ws2.cell(4, ci, h)
        c.font = _f(bold=True, size=8, color=_WHITE)
        c.fill = _fill(_GREEN); c.alignment = _al(h="center"); c.border = _border()
    for col, w in [(1, 16), (2, 14), (3, 14), (4, 14)]:
        ws2.column_dimensions[get_column_letter(col)].width = w

    for r, (m, bud, act) in enumerate(zip(chart_months, exp_budget, exp_actual), 5):
        rf = _fill(_WHITE) if r % 2 == 0 else _fill(_GREY)
        ws2.cell(r, 1, m).fill = rf; ws2.cell(r, 1).border = _border()
        for ci, v in enumerate([bud, act, bud - act], 2):
            c = ws2.cell(r, ci, v)
            c.number_format = "#,##0"; c.alignment = _al(h="right")
            c.fill = rf; c.border = _border()
        ws2.row_dimensions[r].height = 13

    # totals row
    tr = 5 + len(chart_months)
    ws2.cell(tr, 1, "TOTAL").font = _f(bold=True)
    ws2.cell(tr, 1).fill = _fill(_GREEN_L); ws2.cell(tr, 1).border = _border()
    for ci, vals in [(2, exp_budget), (3, exp_actual)]:
        c = ws2.cell(tr, ci, sum(vals))
        c.font = _f(bold=True); c.number_format = "#,##0"
        c.alignment = _al(h="right"); c.fill = _fill(_GREEN_L); c.border = _border()
    c = ws2.cell(tr, 4, sum(exp_budget) - sum(exp_actual))
    c.font = _f(bold=True); c.number_format = "#,##0"
    c.alignment = _al(h="right"); c.fill = _fill(_GREEN_L); c.border = _border()
    ws2.freeze_panes = "A5"

    # ── Sheet 3: Monthly Income ────────────────────────────────────────────
    ws3 = wb.create_sheet("Monthly Income")
    _grid_header(ws3, f"Monthly Income — FY{fiscal_year}",
                 "Godavari Biorefineries Inc", 2)
    for ci, h in enumerate(["Month", "Income Actual"], 1):
        c = ws3.cell(4, ci, h)
        c.font = _f(bold=True, size=8, color=_WHITE)
        c.fill = _fill(_GREEN); c.alignment = _al(h="center"); c.border = _border()
    ws3.column_dimensions["A"].width = 16
    ws3.column_dimensions["B"].width = 16

    for r, (m, inc) in enumerate(zip(chart_months, income_actual), 5):
        rf = _fill(_WHITE) if r % 2 == 0 else _fill(_GREY)
        ws3.cell(r, 1, m).fill = rf; ws3.cell(r, 1).border = _border()
        c = ws3.cell(r, 2, inc)
        c.number_format = "#,##0"; c.alignment = _al(h="right")
        c.fill = rf; c.border = _border()
        ws3.row_dimensions[r].height = 13

    tr3 = 5 + len(chart_months)
    ws3.cell(tr3, 1, "TOTAL").font = _f(bold=True)
    ws3.cell(tr3, 1).fill = _fill(_GREEN_L); ws3.cell(tr3, 1).border = _border()
    c = ws3.cell(tr3, 2, sum(income_actual))
    c.font = _f(bold=True); c.number_format = "#,##0"
    c.alignment = _al(h="right"); c.fill = _fill(_GREEN_L); c.border = _border()
    ws3.freeze_panes = "A5"

    fname = f"GBInc-Analysis-FY{fiscal_year}.xlsx"
    buf = BytesIO(); wb.save(buf)
    return buf.getvalue(), fname
